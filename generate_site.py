from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from pathlib import Path
import calendar, datetime, html, json, re, sys, zlib

try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, OSError):
    pass

SRC = Path(r"C:/Users/garet/OneDrive/桌面/Timetable/ERB Super Timetable 04_checking 11_20260715_V07_HK239HGCW10_REDO.xlsx")
OUTDIR = Path(r"D:/Claude Code/ERB Super Timetable/erb-super-timetable")
OUTDIR.mkdir(parents=True, exist_ok=True)
MONTH_SHEETS = ["June", "July New", "August New", "September New", "October New", "November New", "December New"]
YEAR = 2026
BUILD_ID = "v18l-hk281-l52-helper-20260719a"
CONTEXT_SRC = OUTDIR / "class_context.json"
OVERRIDES_SRC = OUTDIR / "schedule_overrides.json"
VERSIONS_SRC = OUTDIR / "versions.json"
COMPARE_BASELINE = OUTDIR / "versions" / "2026-07-19-V18k"
COMPARE_LABEL = "V18l"
COMPARE_BASELINE_LABEL = "V18k"
EXPECTED_COMPARISON_CHANGES = 62

COURSE_CHINESE_NAMES = {
    "HK239HG": "人工智能知識及應用證書（兼讀制）",
    "HK244EG": "人工智能創作營銷社交媒體內容技巧證書（兼讀制）",
    "HK244HG": "人工智能創作營銷社交媒體內容技巧證書（兼讀制）",
    "HK265HG": "人工智能強化營銷社交媒體內容創作證書（英文授課／兼讀制）",
    "HK280HG": "生成式人工智能商業應用證書（兼讀制）",
    "HK280HS": "生成式人工智能商務應用證書（兼讀制）",
    "HK281DS": "創意數碼媒體設計及製作助理證書",
    "MC0106DS": "創意數碼媒體設計及製作助理證書",
}
COURSE_FAMILIES = [
    {
        "name": "人工智能創作營銷社交媒體內容技巧證書（兼讀制）",
        "members": [
            ("HK244EG", "基督教勵行會", ""),
            ("HK244HG", "基督教勵行會", ""),
            ("HK265HG", "基督教勵行會", "英文授課"),
        ],
    },
    {
        "name": "人工智能知識及應用證書（兼讀制）",
        "members": [("HK239HG", "基督教勵行會", "")],
    },
    {
        "name": "創意數碼媒體設計及製作助理證書",
        "members": [
            ("HK281DS", "基督教勵行會", ""),
            ("MC0106DS", "循道衞理中心", ""),
        ],
    },
    {
        "name": "生成式人工智能商業應用證書（兼讀制）",
        "members": [
            ("HK280HG", "基督教勵行會", ""),
            ("HK280HS", "基督教勵行會", ""),
        ],
    },
]
UPCOMING_AS_OF = datetime.date(2026, 7, 19)
UPCOMING_CLASS_META = {
    "HK280HG · SS": ("基督教勵行會", "上水彩園邨彩湖樓2座地下129舖02室", "CHI"),
    "HK280HS · SS": ("基督教勵行會", "上水彩園邨彩湖樓2座地下129舖02室", "CHI"),
    "HK265HG · FS": ("基督教勵行會", "四海大廈", "ENG"),
    "HK265HG · FS · JUL 2026": ("基督教勵行會", "四海大廈", "ENG"),
    "HK265HG · FS · SEP 2026": ("基督教勵行會", "四海大廈", "ENG"),
    "MC0106DS · 第2班": ("循道衛理中心", "灣仔軒尼詩道22號3樓", "CHI"),
    "HK244HG · CW8": ("基督教勵行會", "彩雲邨", "CHI"),
    "HK244EG · CW": ("基督教勵行會", "九龍彩雲二邨清水灣道55號1樓103室", "CHI"),
    "HK239HG · CW10": ("基督教勵行會", "九龍彩雲二邨清水灣道55號1樓101室", "CHI"),
    "HK281DS · CW7": ("基督教勵行會", "彩雲邨（未有確實街道／房號）", "CHI"),
    "HK239HG · SS": ("基督教勵行會", "上水彩園", "CHI"),
    "HK244EG · FS": ("基督教勵行會", "四海大廈", "CHI"),
    "HK239HG · ST": ("基督教勵行會", "順天", "CHI"),
    "HK239HG · 城巿一條龍": ("基督教勵行會", "彩雲邨", "CHI"),
    "HK239HG · LT": ("基督教勵行會", "藍田", "CHI"),
}
SEN_CODE_RE = re.compile(r"\((PFSA2|QAT7)\)", re.I)

wb = load_workbook(SRC, data_only=False, rich_text=True)
GROUPS = [
    (0, [3,4], 3, "Sunday"),
    (1, [5,6], 6, "Monday"),
    (2, [8,9], 9, "Tuesday"),
    (3, [11,12], 12, "Wednesday"),
    (4, [14,15], 15, "Thursday"),
    (5, [17,18], 18, "Friday"),
    (6, [20,21], 21, "Saturday"),
]
DAY_ROWS = [6,11,16,21,26,31]
MONTH_MAP = {"June":6, "July New":7, "August New":8, "September New":9, "October New":10, "November New":11, "December New":12}

def raw_text(v):
    if v is None:
        return ""
    if isinstance(v, CellRichText):
        return "".join(str(part.text if isinstance(part, TextBlock) else part) for part in v)
    return str(v)

def norm_text(v):
    s = raw_text(v).replace("\r", "\n")
    s = re.sub(r"\n+", " / ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def color_rgb(font):
    col = getattr(font, "color", None)
    if not col:
        return ""
    try:
        if col.type == "rgb" and col.rgb:
            return str(col.rgb).upper()
    except Exception:
        pass
    return ""

def is_red_font(font):
    rgb = color_rgb(font)
    return bool(rgb and rgb.endswith("FF0000"))

def cell_text_runs(cell):
    value = cell.value
    whole_red = is_red_font(cell.font)
    runs = []
    if isinstance(value, CellRichText):
        for part in value:
            if isinstance(part, TextBlock):
                txt = str(part.text)
                red = whole_red or is_red_font(part.font)
            else:
                txt = str(part)
                red = whole_red
            if txt:
                runs.append((txt, red))
    else:
        txt = raw_text(value)
        if txt:
            runs.append((txt, whole_red))
    return normalize_runs(runs)

def normalize_runs(runs):
    out = []
    for txt, red in runs:
        txt = str(txt).replace("\r", "\n")
        txt = re.sub(r"\n+", " / ", txt)
        txt = re.sub(r"\s+", " ", txt)
        if not txt:
            continue
        if out and not out[-1][0].endswith((" ", "/")) and not txt.startswith((" ", "/")):
            out.append((" ", False))
        out.append((txt, red))
    while out and not out[0][0].strip():
        out.pop(0)
    while out and not out[-1][0].strip():
        out.pop()
    if out:
        first_txt, first_red = out[0]
        out[0] = (first_txt.lstrip(), first_red)
        last_txt, last_red = out[-1]
        out[-1] = (last_txt.rstrip(), last_red)
    merged = []
    for txt, red in out:
        if not txt:
            continue
        if merged and merged[-1][1] == red:
            merged[-1] = (merged[-1][0] + txt, red)
        else:
            merged.append((txt, red))
    return merged

def runs_plain(runs):
    return "".join(txt for txt, _ in runs)

def split_runs(runs):
    before, after = [], []
    found = False
    for txt, red in runs:
        if found:
            after.append((txt, red))
            continue
        idx = txt.find(" / ")
        if idx >= 0:
            if txt[:idx].strip():
                before.append((txt[:idx].rstrip(), red))
            rem = txt[idx+3:].lstrip()
            if rem:
                after.append((rem, red))
            found = True
        else:
            before.append((txt, red))
    if not before:
        before = runs[:]
    return before, after

def runs_html(runs):
    bits = []
    for txt, red in runs:
        esc = html.escape(str(txt), quote=True)
        if red:
            bits.append(f'<span class="xl-red">{esc}</span>')
        else:
            bits.append(esc)
    return "".join(bits)

def border_status(cell):
    sides = [getattr(cell.border, edge, None) for edge in ("top", "right", "bottom", "left")]
    styles = [getattr(side, "style", None) for side in sides]
    styles = [s for s in styles if s]
    if not styles:
        return "note"
    dashed = sum("dash" in str(style).lower() for style in styles)
    solid = len(styles) - dashed
    if dashed >= solid:
        return "unconfirmed"
    return "confirmed"

def fill_rgb(cell):
    try:
        fg = cell.fill.fgColor
        if fg.type == "rgb" and fg.rgb:
            return fg.rgb
    except Exception:
        pass
    return ""

def group_for_col(col):
    for idx, cols, day_col, name in GROUPS:
        if col in cols:
            return idx, day_col, name
    return None

def date_for(ws, row, col, month):
    g = group_for_col(col)
    if not g:
        return None
    idx, day_col, name = g
    day_row = max([r for r in DAY_ROWS if r <= row], default=None)
    if day_row is None:
        return None
    day_val = ws.cell(day_row, day_col).value
    if day_val in (None, ""):
        for c in GROUPS[idx][1]:
            v = ws.cell(day_row, c).value
            if isinstance(v, (int, float)) or (isinstance(v, str) and v.strip().isdigit()):
                day_val = v
                break
    if day_val in (None, ""):
        return None
    try:
        day = int(float(day_val))
        return datetime.date(YEAR, month, day)
    except Exception:
        return None

def category(text):
    if "Public Holiday" in text or "假期" in text:
        return "holiday", "Holiday"
    if "YMCA" in text:
        return "ymca", "YMCA SEN"
    if "DGS" in text or "Unreal" in text:
        return "dgs", "DGS / UE"
    if "循道" in text or "MC0106" in text:
        return "methodist", "循道"
    if "勵行" in text or "HK244" in text or "HK239" in text or "HK265" in text or "HK280" in text or "HK281" in text:
        return "erb", "勵行 / ERB"
    if "Mike" in text:
        return "mike", "Mike Sir"
    if "佛教" in text or "Canva" in text:
        return "school", "School"
    return "other", "Other"

def slash_parts(text):
    parts = []
    current = []
    depth = 0
    for char in str(text or ""):
        if char in "([":
            depth += 1
        elif char in ")]" and depth:
            depth -= 1
        if char == "/" and depth == 0:
            value = "".join(current).strip()
            if value:
                parts.append(value)
            current = []
        else:
            current.append(char)
    value = "".join(current).strip()
    if value:
        parts.append(value)
    return parts


def split_title(text):
    parts = slash_parts(text)
    if not parts:
        return text, ""
    return parts[0], " / ".join(parts[1:])

COURSE_CODE_RE = re.compile(r"(?<![A-Z0-9])(?:HK\d+[A-Z]+|MC\d+[A-Z]+|PFSA\d+|QAT\d+|DGS)(?![A-Z0-9])", re.I)
CLASS_RE = re.compile(r"(?<![A-Z0-9])Class\s+([^,/()]+)", re.I)
CODE_PAREN_CLASS_RE = re.compile(r"(?<![A-Z0-9])(?:HK\d+[A-Z]+|MC\d+[A-Z]+|PFSA\d+|QAT\d+)\s*\(([^)]+)\)", re.I)
NAMED_CLASS_RE = re.compile(r"\(([^()]*班)\)", re.I)
LESSON_RE = re.compile(r"(?:^|[\s/\-])L\s*(\d+)(?!\d)", re.I)
TIME_RE = re.compile(r"(?<!\d)([01]?\d|2[0-3])[:：]?([0-5]\d)\s*(am|pm)?\s*(?:-|–|至|to)", re.I)
TIMEISH_RE = re.compile(r"\d{1,2}\s*:?\s*\d{2}|\d{3,4}\s*-|[-–]\s*\d{3,4}")
NAME_WORDS = {"GARETT", "GARRETT", "ANDY", "CALVIN", "MIKE"}


def natural_key(value):
    value = str(value or "").strip().upper()
    parts = re.split(r"(\d+)", value)
    return tuple(int(p) if p.isdigit() else p for p in parts if p != "")


def clean_class(value):
    value = re.sub(r"\s+", " ", str(value or "").strip())
    value = re.sub(r"\s*-\s*L\s*\d+(?!\d).*$", "", value, flags=re.I).strip()
    return value


def parenthetical_class(text):
    text = str(text or "")
    m = CODE_PAREN_CLASS_RE.search(text)
    if not m:
        m = NAMED_CLASS_RE.search(text)
    if not m:
        return ""
    value = clean_class(m.group(1))
    upper = value.upper()
    if not value or upper in NAME_WORDS or TIMEISH_RE.search(value) or len(value) > 12:
        return ""
    return value


def course_sort_parts(text):
    text = str(text or "")
    code_m = COURSE_CODE_RE.search(text)
    code = code_m.group(0).upper() if code_m else ""
    cls = ""
    cls_m = CLASS_RE.search(text)
    if cls_m:
        cls = clean_class(cls_m.group(1))
    if not cls:
        cls = parenthetical_class(text)
    lesson_m = LESSON_RE.search(text)
    lesson = int(lesson_m.group(1)) if lesson_m else 999
    time_m = TIME_RE.search(text)
    if time_m:
        hour = int(time_m.group(1))
        minute = int(time_m.group(2))
        marker = (time_m.group(3) or "").lower()
        if marker == "pm" and hour < 12:
            hour += 12
        elif marker == "am" and hour == 12:
            hour = 0
        start = hour * 60 + minute
    else:
        start = 9999
    return code, cls.upper(), lesson, start


def course_group_label(text, category_label):
    text = str(text or "")
    code, cls, _lesson, _start = course_sort_parts(text)
    if code and not cls:
        cls = INFERRED_CLASS_BY_CODE.get(code, "")
    if code:
        return f"{code} · {cls}" if cls else code
    if "Mike Sir" in text:
        return "Mike Sir"
    return category_label or "Other"


def event_sort_key(ev):
    code, cls, lesson, start = course_sort_parts(ev.get("text", ""))
    return (start, 0 if code else 1, natural_key(code), natural_key(cls), lesson, ev.get("row", 999), ev.get("col", 999), ev.get("text", ""))


LESSON_NUMBER_RE = re.compile(r"\bL\s*(\d+)\b", re.I)


def span_lesson_number(ev):
    match = LESSON_NUMBER_RE.search(str(ev.get("text") or ""))
    return int(match.group(1)) if match else None


def split_span_instances(events):
    """Split repeated cohorts when the lesson sequence restarts at L1/L2."""
    events_by_date = {}
    for event in sorted(events, key=lambda item: (item["date"], event_sort_key(item))):
        events_by_date.setdefault(event["date"], []).append(event)

    instances = []
    current = []
    previous_max_lesson = None
    for _date, day_events in sorted(events_by_date.items()):
        lesson_numbers = [number for number in (span_lesson_number(event) for event in day_events) if number is not None]
        day_min = min(lesson_numbers) if lesson_numbers else None
        starts_new_cohort = (
            bool(current)
            and day_min is not None
            and previous_max_lesson is not None
            and day_min <= 2
            and day_min < previous_max_lesson
        )
        if starts_new_cohort:
            instances.append(current)
            current = []
            previous_max_lesson = None
        current.extend(day_events)
        if lesson_numbers:
            previous_max_lesson = max(previous_max_lesson or 0, max(lesson_numbers))
    if current:
        instances.append(current)
    return instances


events = []
context_events = []
by_date = {}
INFERRED_CLASS_BY_CODE = {}

for sheet in MONTH_SHEETS:
    ws = wb[sheet]
    month = MONTH_MAP[sheet]
    seen = set()
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            text_runs = cell_text_runs(cell)
            text = runs_plain(text_runs)
            if not text:
                continue
            if row in (1,2,3,4,5):
                continue
            if row in DAY_ROWS and text.isdigit():
                continue
            dt = date_for(ws, row, col, month)
            if not dt:
                continue
            key = (dt.isoformat(), row, text)
            if key in seen:
                continue
            seen.add(key)
            status = border_status(cell)
            cat, cat_label = category(text)
            title_runs, detail_runs = split_runs(text_runs)
            title, detail = runs_plain(title_runs), runs_plain(detail_runs)
            ev = {
                "date": dt.isoformat(), "month": sheet, "row": row, "col": col, "cell": cell.coordinate,
                "text": text, "title": title, "detail": detail, "status": status, "fill": fill_rgb(cell),
                "category": cat, "category_label": cat_label,
                "html": runs_html(text_runs), "title_html": runs_html(title_runs), "detail_html": runs_html(detail_runs),
                "red": any(red for _, red in text_runs),
            }
            events.append(ev)
            by_date.setdefault(dt.isoformat(), []).append(ev)

override_revision = ""
override_confirmation = ""
if OVERRIDES_SRC.exists():
    override_data = json.loads(OVERRIDES_SRC.read_text(encoding="utf-8"))
    override_revision = str(override_data.get("revision", ""))
    override_confirmation = str(override_data.get("confirmation", ""))
    override_default_status = override_data.get("default_status")
    if override_default_status not in {None, "confirmed", "unconfirmed", "note"}:
        raise ValueError(f"Invalid override default status: {override_default_status}")
    override_index = {}
    override_cell_index = {}
    for event in events:
        cell_key = (event["date"], event["cell"].upper())
        if cell_key in override_cell_index:
            raise ValueError(f"Duplicate workbook cell override target: {cell_key}")
        override_cell_index[cell_key] = event
        code, cls, lesson, _start = course_sort_parts(event["text"])
        if code and cls:
            key = (event["date"], code.upper(), cls.upper(), lesson)
            if key in override_index:
                raise ValueError(f"Duplicate workbook override target: {key}")
            override_index[key] = event

    for index, item in enumerate(override_data.get("overrides", []), 1):
        match_cell = str(item.get("match_cell", "")).strip().upper()
        if match_cell:
            key = (item["date"], match_cell)
            event = override_cell_index.get(key)
        else:
            match_lesson = item.get("match_lesson", item.get("lesson"))
            key = (
                item["date"],
                str(item.get("match_course_code", item["course_code"])).upper(),
                str(item.get("match_class", item["class"])).upper(),
                999 if match_lesson is None else int(match_lesson),
            )
            event = override_index.get(key)
        if event is None:
            raise ValueError(f"Override entry {index} has no workbook lesson: {key}")
        text = norm_text(item["text"])
        code, cls, lesson, _start = course_sort_parts(text)
        output_lesson = item.get("lesson")
        output_key = (
            item["date"],
            str(item["course_code"]).upper(),
            str(item["class"]).upper(),
            999 if output_lesson is None else int(output_lesson),
        )
        if (event["date"], code.upper(), cls.upper(), lesson) != output_key:
            raise ValueError(f"Override entry {index} text does not match its output key: {output_key}")
        title, detail = split_title(text)
        cat, cat_label = category(text)
        status = item.get("status", override_default_status or event["status"])
        if status not in {"confirmed", "unconfirmed", "note"}:
            raise ValueError(f"Override entry {index} has invalid status: {status}")
        event.update({
            "text": text,
            "title": title,
            "detail": detail,
            "category": cat,
            "category_label": cat_label,
            "html": html.escape(text, quote=True),
            "title_html": html.escape(title, quote=True),
            "detail_html": html.escape(detail, quote=True),
            "red": bool(item.get("red", False)),
            "status": status,
            "teacher": item.get("teacher", ""),
            "source": item.get("source", ""),
            "excluded": bool(item.get("exclude", False)),
        })

events = [event for event in events if not event.get("excluded")]
by_date = {}
for event in events:
    by_date.setdefault(event["date"], []).append(event)

if CONTEXT_SRC.exists():
    raw_context = json.loads(CONTEXT_SRC.read_text(encoding="utf-8"))
    baseline_keys = {
        (event["date"],) + course_sort_parts(event["text"])[:3]
        for event in events
    }
    for index, item in enumerate(raw_context, 1):
        dt = datetime.date.fromisoformat(item["date"])
        text = norm_text(item["text"])
        cat, cat_label = category(text)
        if cat not in {"erb", "methodist"}:
            raise ValueError(f"Context entry {index} must be an ERB or Methodist lesson")
        key = (dt.isoformat(),) + course_sort_parts(text)[:3]
        if key in baseline_keys:
            raise ValueError(f"Context entry {index} duplicates a workbook lesson: {key}")
        title, detail = split_title(text)
        teacher = str(item.get("teacher", "Other tutor / TBC")).strip()
        is_garett_teacher = bool(re.search(r"\bGar(?:e|r)tt\b", teacher, re.I))
        layer = item.get("layer", "mine" if is_garett_teacher else "class")
        if layer not in {"mine", "class"}:
            raise ValueError(f"Context entry {index} has invalid layer: {layer}")
        if is_garett_teacher and layer != "mine":
            raise ValueError(
                f"Context entry {index} assigns Garett to the hidden class layer: {text}"
            )
        ev = {
            "date": dt.isoformat(), "month": calendar.month_name[dt.month], "row": 999, "col": 999,
            "cell": f"context-{index}", "text": text, "title": title, "detail": detail,
            "status": item.get("status", "unconfirmed"), "fill": "", "category": cat,
            "category_label": cat_label, "html": html.escape(text, quote=True),
            "title_html": html.escape(title, quote=True), "detail_html": html.escape(detail, quote=True),
            "red": bool(item.get("red", False)), "layer": layer,
            "teacher": teacher, "helper": str(item.get("helper", "")).strip(),
            "source": item.get("source", ""),
        }
        context_events.append(ev)
        by_date.setdefault(dt.isoformat(), []).append(ev)

display_events = events + context_events

baseline_events = json.loads((COMPARE_BASELINE / "events.json").read_text(encoding="utf-8"))
baseline_context = json.loads((COMPARE_BASELINE / "class_context.json").read_text(encoding="utf-8"))

def workbook_identity(item):
    return tuple(item.get(key) for key in ("date", "month", "row", "col", "cell"))

def context_identity(item):
    return item.get("date"), item.get("source", "")

baseline_event_map = {workbook_identity(item): item for item in baseline_events}
baseline_context_map = {context_identity(item): item for item in baseline_context}

def effective_context_layer(item):
    # Published snapshots before V15 used class as the implicit context layer.
    return item.get("layer", "class")

for current in events:
    previous = baseline_event_map.get(workbook_identity(current))
    changed = previous is None or any(current.get(key) != previous.get(key) for key in ("text", "status"))
    current["changed_in_version"] = changed
    current["change_kind"] = "new" if previous is None else "changed" if changed else ""
    current["previous_text"] = previous.get("text", "") if previous else ""
    current["previous_status"] = previous.get("status", "") if previous else ""

for current in context_events:
    previous = baseline_context_map.get(context_identity(current))
    changed = (
        previous is None
        or any(current.get(key) != previous.get(key) for key in ("text", "status", "teacher"))
        or current.get("layer") != effective_context_layer(previous)
    )
    current["changed_in_version"] = changed
    current["change_kind"] = "new" if previous is None else "changed" if changed else ""
    current["previous_text"] = previous.get("text", "") if previous else ""
    current["previous_status"] = previous.get("status", "") if previous else ""

changed_events = [event for event in display_events if event.get("changed_in_version")]
if len(changed_events) != EXPECTED_COMPARISON_CHANGES:
    raise ValueError(
        f"Expected {EXPECTED_COMPARISON_CHANGES} changes in {COMPARE_LABEL}, found {len(changed_events)}"
    )

comparison_legend_html = (
    f'<div class="legend-card"><span class="sample changed-sample"></span> Changed in {COMPARE_LABEL}</div>'
    if changed_events else ""
)
comparison_filter_html = (
    f'<button class="filter course-filter change-filter" data-filter="changed" '
    f'data-first-date="{min(event["date"] for event in changed_events)}">'
    f'Changed in {COMPARE_LABEL} ({len(changed_events)})</button>'
    if changed_events else ""
)

for ds in by_date:
    by_date[ds].sort(key=event_sort_key)

classes_by_code = {}
for event in display_events:
    code, cls, _lesson, _start = course_sort_parts(event["text"])
    if code and cls:
        classes_by_code.setdefault(code, set()).add(cls)
INFERRED_CLASS_BY_CODE = {
    code: next(iter(classes))
    for code, classes in classes_by_code.items()
    if len(classes) == 1
}

_events_by_base_group = {}
for event in display_events:
    base_label = course_group_label(event["text"], event["category_label"])
    _events_by_base_group.setdefault(base_label, []).append(event)

for base_label, group_events in _events_by_base_group.items():
    instances = split_span_instances(group_events)
    if len(instances) == 1:
        for event in instances[0]:
            event["group_label"] = base_label
        continue
    for instance in instances:
        first_day = datetime.date.fromisoformat(min(event["date"] for event in instance))
        cohort_label = f"{base_label} · {calendar.month_abbr[first_day.month].upper()} {first_day.year}"
        for event in instance:
            event["group_label"] = cohort_label

_hk265_cohorts = {
    event["group_label"]: [item for item in display_events if item["group_label"] == event["group_label"]]
    for event in display_events
    if event["group_label"].startswith("HK265HG · FS · ")
}
_hk265_signature = sorted(
    (label, min(item["date"] for item in items), len(items))
    for label, items in _hk265_cohorts.items()
)
if _hk265_signature != [
    ("HK265HG · FS · JUL 2026", "2026-07-24", 12),
    ("HK265HG · FS · SEP 2026", "2026-09-16", 12),
]:
    raise ValueError(f"HK265HG FS cohort split mismatch: {_hk265_signature}")

_group_labels = sorted({e["group_label"] for e in display_events}, key=lambda label: (0 if COURSE_CODE_RE.fullmatch(label.split(" · ", 1)[0]) or label == "DGS" else 1, natural_key(label.split(" · ", 1)[0]), natural_key(label.split(" · ", 1)[1] if " · " in label else ""), label))

# Preserve existing filter IDs across releases. A newly introduced course must not
# renumber unrelated workbook events or invalidate the immutable baseline hash.
_baseline_group_slugs = {
    item["group_label"]: item["group"]
    for item in baseline_events
    if item.get("group_label") and item.get("group")
}
_baseline_index = (COMPARE_BASELINE / "index.html").read_text(encoding="utf-8")
for match in re.finditer(r'data-group="([^"]+)" data-group-label="([^"]+)"', _baseline_index):
    slug, label = html.unescape(match.group(1)), html.unescape(match.group(2))
    _baseline_group_slugs.setdefault(label, slug)

_group_slugs = {}
_used_group_slugs = set()
for label in _group_labels:
    slug = _baseline_group_slugs.get(label)
    if slug and slug not in _used_group_slugs:
        _group_slugs[label] = slug
        _used_group_slugs.add(slug)
_next_group_number = max(
    [int(match.group(1)) for slug in _used_group_slugs if (match := re.fullmatch(r"g(\d+)", slug))]
    or [0]
) + 1
for label in _group_labels:
    if label in _group_slugs:
        continue
    while f"g{_next_group_number:02d}" in _used_group_slugs:
        _next_group_number += 1
    slug = f"g{_next_group_number:02d}"
    _group_slugs[label] = slug
    _used_group_slugs.add(slug)
    _next_group_number += 1
for ev in display_events:
    ev["group"] = _group_slugs[ev["group_label"]]

def ehtml(s):
    return html.escape(str(s or ""), quote=True)

GROUPS = []

def is_personal_assignment(event):
    if event.get("layer") == "class":
        return False
    category = event.get("category")
    if category in {"ymca", "dgs"}:
        return True
    teacher = str(event.get("teacher", "")).strip()
    text = str(event.get("text", ""))
    if "GARETT NOT REQUIRED" in text.upper():
        return False
    return bool(re.search(r"\bGar(?:e|r)tt\b", teacher or text, re.I))


def lifecycle_status(group_events):
    statuses = {event["status"] for event in group_events}
    if statuses == {"note"}:
        return "note"
    future_events = [
        event for event in group_events
        if event["date"] >= UPCOMING_AS_OF.isoformat()
    ]
    if any(
        event["status"] == "unconfirmed" and is_personal_assignment(event)
        for event in future_events
    ):
        return "pending"
    if not future_events:
        return "completed"
    if any(is_personal_assignment(event) for event in future_events):
        return "upcoming"
    return "context"


for label in _group_labels:
    slug = _group_slugs[label]
    group_events = [e for e in display_events if e["group"] == slug]
    group_status = lifecycle_status(group_events)
    first_date = min(e["date"] for e in group_events)
    GROUPS.append((label, slug, group_status, first_date))

CSS = r'''
*{box-sizing:border-box}html{-webkit-text-size-adjust:100%;scroll-behavior:smooth}body{margin:0;font-family:"Segoe UI Variable","Segoe UI",-apple-system,BlinkMacSystemFont,Roboto,"Noto Sans TC","Microsoft JhengHei",Arial,sans-serif;background:#eef1f6;color:#1d2734;line-height:1.42}.xl-red{color:#d60000;font-weight:700}.chip .xl-red{color:#d60000}.modal-body .xl-red{color:#d60000;font-weight:750}.wrap{max-width:1280px;margin:0 auto;padding:28px 16px 70px}.hero{display:flex;justify-content:space-between;gap:16px;align-items:flex-start}.title{font-size:30px;font-weight:850;letter-spacing:-.45px;margin:0}.title .y{color:#0f7d7d}.sub{color:#64707f;margin:6px 0 0;font-size:14px}.actions{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}.btn{border:1px solid #d8e0ea;background:#fff;color:#344153;border-radius:10px;padding:8px 12px;font-weight:750;font-size:13px;text-decoration:none;box-shadow:0 1px 2px rgba(20,30,50,.05)}.btn:hover{border-color:#8bb8bd}.stats,.legend{display:flex;gap:10px;flex-wrap:wrap;margin:18px 0 6px}.stat,.legend-card{background:#fff;border:1px solid #e2e7ef;border-radius:12px;padding:8px 13px;font-size:13px;color:#46505e;box-shadow:0 1px 2px rgba(20,30,50,.05)}.stat b{color:#1d2734;font-size:15px}.legend-card{display:flex;align-items:center;gap:8px}.sample{width:28px;height:18px;border-radius:6px;background:#f9fcff}.sample.confirmed{border:3px solid #1d2734}.sample.unconfirmed{border:3px dashed #1d2734}.sample.note{border:2px solid #b8c1ce}.filters{display:flex;gap:8px;flex-wrap:wrap;margin:16px 0}.filter{border:2px solid #dbe3ed;background:#fff;border-radius:999px;padding:7px 11px;font-weight:750;font-size:12px;color:#4c5a6b;cursor:pointer}.filter.confirmed{border:3px solid #1d2734}.filter.unconfirmed{border:3px dashed #1d2734}.filter.mixed{border:3px dashed #1d2734;box-shadow:inset 0 0 0 2px rgba(29,39,52,.18),0 1px 2px rgba(20,30,50,.05)}.filter.note{border:2px solid #b8c1ce;color:#69737f}.filter.active{background:#0f7d7d;color:#fff;border-color:#1d2734}.filter.active.confirmed{border-style:solid}.filter.active.unconfirmed,.filter.active.mixed{border-style:dashed}.filter.active.note{border-color:#b8c1ce}.section-h{font-size:13px;text-transform:uppercase;letter-spacing:.8px;color:#8a94a2;font-weight:800;margin:24px 2px 10px}.month{background:#fff;border:1px solid #e2e7ef;border-radius:16px;padding:16px;margin-top:18px;box-shadow:0 1px 3px rgba(20,30,50,.06)}.month h2{margin:0 0 12px;font-size:20px;font-weight:850;letter-spacing:-.2px}.gridwrap{overflow-x:auto}.grid{display:grid;grid-template-columns:repeat(7,minmax(124px,1fr));gap:7px;min-width:868px}.dow{font-size:11.5px;font-weight:800;color:#98a2af;text-align:center;padding:2px 0;text-transform:uppercase;letter-spacing:.5px}.cell{min-height:132px;border:1px solid #e8ecf3;border-radius:10px;padding:6px;background:#fcfdff;display:flex;flex-direction:column;gap:4px}.cell.out{background:#f4f6f9;border-style:dashed;opacity:.5}.cell.wknd{background:#f7f9fc}.cell.today{outline:3px solid #0f7d7d;outline-offset:1px}.dnum{font-size:12px;font-weight:850;color:#9aa4b1}.cell.has .dnum{color:#2d3948}.dnum{display:flex;align-items:baseline;gap:3px;white-space:nowrap}.dnum .dmon{font-size:13px;font-weight:900;color:#667384;text-transform:uppercase}.dnum .dday{font-size:12px;font-weight:900;color:inherit}.dnum .dweekday{margin-left:auto;font-size:10.5px;font-weight:850;color:#8792a0;text-transform:uppercase}.chip{border-radius:8px;padding:5px 7px 6px;background:#fff;box-shadow:0 1px 1px rgba(20,30,50,.04);cursor:pointer;overflow:visible}.chip.confirmed{border-style:solid!important;border-width:2.5px!important;border-color:#1d2734!important;box-shadow:0 0 0 1px rgba(29,39,52,.10) inset,0 1px 1px rgba(20,30,50,.04)}.chip.unconfirmed{border-style:dashed!important;border-width:2.5px!important;border-color:#1d2734!important;box-shadow:0 0 0 1px rgba(29,39,52,.10) inset,0 1px 1px rgba(20,30,50,.04)}.chip.note{border-style:solid!important;border-width:1.5px!important;border-color:#9aa4b2!important;background:#f8fafc}.chip .top{display:flex;justify-content:space-between;gap:6px;align-items:flex-start}.chip .cat{font-size:10px;font-weight:850;text-transform:uppercase;letter-spacing:.35px;opacity:.82}.chip .status{font-size:9.5px;font-weight:850;white-space:nowrap}.class-id{display:inline-flex;align-items:center;align-self:flex-start;gap:4px;max-width:100%;margin-top:3px;padding:2px 5px;border:1.5px solid hsl(var(--class-hue),72%,14%);border-radius:4px;background:hsl(var(--class-hue),72%,22%);color:#fff;font-size:9.3px;font-weight:900;line-height:1.15;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;box-shadow:0 1px 1px rgba(12,18,28,.22)}.class-dot{width:7px;height:7px;flex:0 0 7px;border-radius:50%;background:#fff;box-shadow:0 0 0 1px rgba(255,255,255,.45)}.chip .ttl{font-size:11.5px;font-weight:850;margin-top:2px;color:#172232;line-height:1.22}.chip .det{font-size:10.2px;color:#596676;margin-top:2px;line-height:1.22;display:block;white-space:normal;overflow:visible}.chip .fulltxt{display:none}.cat-ymca{background:#e3f7fa}.cat-erb{background:#fff1e6}.cat-methodist{background:#eef0ff}.cat-dgs{background:#ecfdf3}.cat-holiday{background:#f3f4f6}.cat-mike{background:#fff8db}.cat-school{background:#fce7f3}.cat-other{background:#f6f7fb}.agenda{display:none}.aday{display:flex;gap:10px;align-items:flex-start;padding:9px 2px;border-top:1px solid #eef1f6}.aday.empty-day{display:none}.aday:first-child{border-top:none}.adate{flex:0 0 48px;text-align:center}.adate .adow{display:block;font-size:11px;font-weight:800;color:#98a2af;text-transform:uppercase}.adate .amon{display:block;font-size:10px;font-weight:900;color:#0f7d7d;text-transform:uppercase;letter-spacing:.05em;line-height:1}.adate .anum{display:block;font-size:20px;font-weight:850;color:#3a4452;line-height:1.05}.achips{flex:1;min-width:0;display:flex;flex-direction:column;gap:6px}.foot{margin-top:28px;color:#7a8492;font-size:12.5px;border-top:1px solid #e2e7ef;padding-top:14px}.modal{position:fixed;inset:0;background:rgba(18,26,38,.56);display:flex;align-items:center;justify-content:center;padding:18px;z-index:50}.modal[hidden]{display:none}.modal-card{background:#fff;border-radius:16px;max-width:560px;width:100%;padding:20px 20px 22px;box-shadow:0 14px 44px rgba(15,25,45,.32);position:relative;max-height:88vh;overflow:auto}.modal-x{position:absolute;top:8px;right:12px;border:none;background:transparent;font-size:26px;line-height:1;color:#98a2af;cursor:pointer}.modal-h{font-size:20px;font-weight:850;padding-right:24px}.modal-date{color:#69737f;font-size:13px;margin-top:2px}.modal-body{white-space:pre-wrap;margin-top:14px;font-size:15px}.pill{display:inline-block;border-radius:999px;padding:4px 9px;font-size:12px;font-weight:850;margin-top:10px;margin-right:6px}.pill.confirmed{background:#e7f6ee;color:#16623d}.pill.unconfirmed{background:#fff3df;color:#a25600}.pill.note{background:#eef2f7;color:#596676}@media (max-width:820px){.wrap{padding:14px 10px 48px}.hero{display:block}.title{font-size:22px}.sub{font-size:13px}.actions{justify-content:flex-start;margin-top:10px}.month{padding:12px 10px 14px}.month h2{font-size:17px}.stats,.legend{gap:7px}.stat,.legend-card{font-size:12px;padding:7px 10px}}@media (orientation:portrait) and (max-width:820px){.gridwrap{display:none}.agenda{display:block}.aday{scroll-margin-top:18px}.aday.today{background:linear-gradient(90deg,rgba(15,125,125,.10),transparent);border-radius:14px}}@media (orientation:landscape) and (max-height:540px){.sub,.stats,.legend,.filters,.foot{display:none}.month{padding:8px;margin-top:10px}.month h2{font-size:15px;margin:0 0 6px}.gridwrap{overflow:visible}.grid{min-width:0;grid-template-columns:repeat(7,minmax(0,1fr));gap:3px}.dow{font-size:8.5px;letter-spacing:0;padding:0;text-transform:uppercase}.cell{min-height:98px;height:auto;padding:2px;border-radius:5px;overflow:visible}.dnum{font-size:8.5px;gap:2px}.dnum .dmon{font-size:9px}.dnum .dday{font-size:8.5px}.dnum .dweekday{font-size:6.8px}.chip{padding:2px 3px 3px;border-radius:4px;overflow:visible}.chip.confirmed{border-width:1.8px!important}.chip.unconfirmed{border-width:1.8px!important}.chip.note{border-width:1.3px!important}.chip .top{margin-bottom:1px}.chip .cat,.chip .status{font-size:5.8px;font-weight:550;letter-spacing:0}.class-id{margin-top:1px;padding:1px 2px;gap:2px;font-size:5.8px;border-radius:2px}.class-dot{width:4px;height:4px;flex-basis:4px}.chip .ttl,.chip .det{display:none}.chip .fulltxt{display:block;font-size:6.6px;font-weight:400;line-height:1.14;color:#172232;white-space:normal;overflow:visible;word-break:break-word;overflow-wrap:anywhere}}@media print{body{background:#fff}.month,.stat,.legend-card{box-shadow:none}.actions,.filters{display:none}.gridwrap{overflow:visible}.grid{min-width:0}.chip{-webkit-print-color-adjust:exact;print-color-adjust:exact}}
'''

CSS += r'''
.foot{overflow-wrap:anywhere}
.filter.upcoming{border:2.5px solid #0f7074;background:#e8f7f5;color:#0b5f62}
.filter.pending{border:3px dashed #b65a00;background:#fff1d7;color:#8a4200}
.filter.completed{border:2px solid #8a96a5;background:#edf1f5;color:#566273}
.filter.context{border:2px solid #6f60aa;background:#f0eef9;color:#51458a}
.filter.active.upcoming{border-style:solid;border-color:#09565a;background:#0f7074;color:#fff}
.filter.active.pending{border-style:dashed;border-color:#7f3f00;background:#a94f00;color:#fff}
.filter.active.completed{border-style:solid;border-color:#566273;background:#667384;color:#fff}
.filter.active.context{border-style:solid;border-color:#596676;background:#596676;color:#fff}
.filter-status-summary{display:flex;align-items:center;gap:7px;flex-wrap:wrap;margin:8px 0 3px}.filter-status-total{color:#425064;font-size:11px;font-weight:900}.filter-status-pill{display:inline-flex;align-items:center;gap:5px;min-height:25px;padding:3px 7px;border:1px solid #cfd8e5;border-radius:5px;background:#fff;color:#526073;font-size:10px;font-weight:850}.filter-status-swatch{width:12px;height:12px;border-radius:3px}.filter-status-swatch.upcoming{border:2px solid #0f7074;background:#e8f7f5}.filter-status-swatch.pending{border:2px dashed #b65a00;background:#fff1d7}.filter-status-swatch.completed{border:2px solid #8a96a5;background:#edf1f5}.filter-status-swatch.context{border:2px solid #6f60aa;background:#f0eef9}
@media (min-width:821px){.wrap{width:100%;max-width:none;margin:0;padding:28px clamp(16px,1.6vw,36px) 70px}}
.title,.month h2{letter-spacing:0}
.layer-controls{margin:16px 0 2px}.layer-controls .section-h{margin:0 0 8px}.layer-switch{display:inline-grid;grid-template-columns:repeat(3,minmax(104px,1fr));padding:3px;border:1px solid #cfd8e5;border-radius:8px;background:#e5eaf1;box-shadow:0 1px 2px rgba(20,30,50,.06)}.mode-filter{min-height:34px;border:0;border-radius:6px;padding:6px 11px;background:transparent;color:#4c5a6b;font:inherit;font-size:12px;font-weight:800;cursor:pointer}.mode-filter.active{background:#fff;color:#145f63;box-shadow:0 1px 3px rgba(20,30,50,.18)}.sample.class-layer{border:2px solid #8b80aa;background:#fff1e6;box-shadow:inset 4px 0 0 #a99bc7}.chip.layer-class.confirmed{box-shadow:inset 3px 0 0 #a99bc7,0 0 0 1px rgba(29,39,52,.10),0 1px 1px rgba(20,30,50,.04)}.chip.layer-class.unconfirmed{box-shadow:inset 3px 0 0 #a99bc7,0 0 0 1px rgba(29,39,52,.10),0 1px 1px rgba(20,30,50,.04)}.modal-source{margin-top:14px;padding-top:10px;border-top:1px solid #e4e9f0;color:#6c7786;font-size:12px}.pill.class-layer{background:#eeeaf7;color:#5d537a}
.chip.erb-compact{position:relative;display:grid;justify-items:center;gap:2px;padding:6px 7px 7px;text-align:center}
.chip.erb-compact .status{position:absolute;top:4px;right:6px;z-index:1}
.chip.erb-compact .class-id{display:inline-flex;align-items:center;align-self:center;justify-content:center;gap:4px;max-width:calc(100% - 24px);min-height:18px;margin:0;padding:2px 6px;border:2px solid hsl(var(--class-hue),72%,38%);border-radius:4px;background:#fff;color:hsl(var(--class-hue),72%,24%);font-size:9.3px;font-weight:900;line-height:1.1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;box-shadow:0 1px 2px rgba(12,18,28,.12);cursor:pointer}
.chip.erb-compact .class-id:hover,.chip.erb-compact .class-id:focus-visible{background:hsl(var(--class-hue),72%,94%);outline:2px solid #ffc857;outline-offset:1px}.chip.erb-compact .class-id.card-filter-active{background:hsl(var(--class-hue),72%,38%);color:#fff}.chip.erb-compact .class-id.card-filter-active .class-dot{background:#fff}
.chip.erb-compact .class-dot{width:7px;height:7px;flex:0 0 7px;border-radius:50%;background:hsl(var(--class-hue),72%,38%);box-shadow:none}
.erb-meta,.erb-course,.erb-foot{width:100%;text-align:center;overflow-wrap:anywhere}
.erb-meta{font-size:10px;font-weight:850;line-height:1.15;color:#172232}
.erb-teacher{color:#0f6868}
.erb-teacher.is-alert{color:#d60000}
.erb-course{font-size:10.3px;font-weight:750;line-height:1.17;color:#263343}
.erb-foot{font-size:9.9px;font-weight:800;line-height:1.15;color:#4d5b6c;font-variant-numeric:tabular-nums}
.erb-sep{padding:0 3px;color:#8d97a4}
.card-note{color:#d60000;font-weight:850}
.chip.erb-compact .fulltxt{display:none!important}
@media (orientation:landscape) and (max-height:540px){.chip.erb-compact{gap:1px;padding:3px 4px 4px}.chip.erb-compact .class-id{min-height:10px;max-width:calc(100% - 14px);padding:1px 3px;font-size:6px;border-width:1.3px;border-radius:2px}.chip.erb-compact .class-dot{width:4px;height:4px;flex-basis:4px}.erb-meta,.erb-course,.erb-foot{font-size:6px;line-height:1.08}.chip.erb-compact .status{top:2px;right:3px;font-size:5.8px}}
@media (max-width:520px){.layer-switch{display:grid;width:100%;grid-template-columns:repeat(3,minmax(0,1fr))}.mode-filter{padding:6px 4px}}
@media (orientation:landscape) and (max-height:540px){.layer-controls{display:none}}
'''

CSS += r'''
.course-code-legend{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:8px;margin:0 0 8px}.course-family-card{min-width:0;padding:9px 10px;border:1px solid #dbe2ec;border-left:5px solid #0f7d7d;border-radius:6px;background:#fff}.course-family-name{color:#1d2734;font-size:12px;font-weight:900;line-height:1.28}.course-family-members{display:flex;gap:5px;flex-wrap:wrap;margin-top:7px}.course-family-member{display:inline-flex;align-items:center;gap:5px;min-height:24px;padding:3px 6px;border:1px solid #cbd5e1;border-radius:4px;background:#f7fafc;color:#526174;font-size:9.5px;font-weight:750;line-height:1.15}.course-family-member b{color:#1d2734;white-space:nowrap}.course-family-member em{color:#8a4b00;font-style:normal;font-weight:850}.sample.class-layer{box-shadow:inset 8px 0 0 #8c78b5}.chip.layer-class{padding-left:14px}.chip.layer-class.confirmed{box-shadow:inset 8px 0 0 #8c78b5,0 0 0 1px rgba(29,39,52,.10),0 1px 1px rgba(20,30,50,.04)}.chip.layer-class.unconfirmed{box-shadow:inset 8px 0 0 #8c78b5,0 0 0 1px rgba(29,39,52,.10),0 1px 1px rgba(20,30,50,.04)}.chip.layer-class.note{box-shadow:inset 8px 0 0 #8c78b5,0 1px 1px rgba(20,30,50,.04)}.chip.erb-compact.layer-class{padding-left:14px}.context-teacher{font-weight:850;color:#6c568f}.overlap-group{position:relative;display:flex;flex-direction:column;gap:4px;min-width:0}.overlap-group.overlap-active{padding:3px 12px 3px 0}.overlap-group.overlap-active::after{content:"";position:absolute;top:1px;right:1px;bottom:1px;width:8px;border:2.5px solid #536170;border-left:0;border-radius:0 5px 5px 0;pointer-events:none}.overlap-group.overlap-active>.chip.layer-mine:first-child{outline:2px solid rgba(15,125,125,.20);outline-offset:1px}
@media (max-width:820px){.course-code-legend{grid-template-columns:1fr}.course-family-name{font-size:11.5px}.course-family-member{font-size:9px}.overlap-group.overlap-active{padding-right:13px}}
@media (orientation:landscape) and (max-height:540px){.course-code-legend{display:none}.chip.layer-class,.chip.erb-compact.layer-class{padding-left:7px}.chip.layer-class.confirmed,.chip.layer-class.unconfirmed,.chip.layer-class.note{box-shadow:inset 4px 0 0 #8c78b5,0 1px 1px rgba(20,30,50,.04)}.overlap-group{gap:2px}.overlap-group.overlap-active{padding:1px 6px 1px 0}.overlap-group.overlap-active::after{width:4px;border-width:1.4px}}
'''

CSS += r'''
.day-hours{display:inline-flex;align-items:center;justify-content:center;min-width:28px;padding:1px 5px;border:1px solid #78b8bb;border-radius:999px;background:#e4f6f5;color:#0b6669;font-size:9px;font-weight:900;line-height:1.15;white-space:nowrap;font-variant-numeric:tabular-nums;box-shadow:0 1px 1px rgba(20,30,50,.08)}.day-hours[hidden]{display:none!important}.dnum .day-hours{margin-left:auto}.dnum .day-hours:not([hidden])+.dweekday{margin-left:6px}.dnum .day-hours[hidden]+.dweekday{margin-left:auto}.adate .day-hours{margin:4px auto 0;padding-inline:4px;font-size:8.5px}
.chip.layer-class{--context-color:#8c78b5}.chip.cat-erb.layer-class{--context-color:#d9772e}.chip.cat-methodist.layer-class{--context-color:#7567b9}.chip.cat-ymca.layer-class{--context-color:#238a99}.chip.cat-dgs.layer-class{--context-color:#31865c}.sample.class-layer{box-shadow:inset 7px 0 0 #d9772e,inset -7px 0 0 #7567b9}.chip.layer-class.confirmed{box-shadow:inset 8px 0 0 var(--context-color),0 0 0 1px rgba(29,39,52,.10),0 1px 1px rgba(20,30,50,.04)}.chip.layer-class.unconfirmed{box-shadow:inset 8px 0 0 var(--context-color),0 0 0 1px rgba(29,39,52,.10),0 1px 1px rgba(20,30,50,.04)}.chip.layer-class.note{box-shadow:inset 8px 0 0 var(--context-color),0 1px 1px rgba(20,30,50,.04)}.floating-mode-switch{position:fixed;z-index:40;left:clamp(58px,3.6vw,138px);bottom:24px;display:grid;grid-template-columns:repeat(3,48px);gap:3px;padding:4px;border:2px solid #fff;border-radius:12px;background:#dfe6ec;box-shadow:0 5px 18px rgba(25,38,55,.30);font-family:inherit}.mode-option,.today-option{width:48px;height:44px;border:0;border-radius:8px;background:#fff;color:#405064;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:2px;cursor:pointer;font-family:inherit}.mode-option:hover,.today-option:hover{background:#eef7f7}.mode-option:focus-visible,.today-option:focus-visible{outline:3px solid #ffc857;outline-offset:1px}.mode-option.active{background:#0f7074;color:#fff}.today-option{color:#0f7074}.today-option .mode-main{font-size:8px}.mode-main{font-size:12px;font-weight:900;line-height:1}.mode-sub{font-size:7px;font-weight:850;line-height:1;text-transform:uppercase}.grid .holiday-cell{background:#f1f3f6}.grid .holiday-cell .chip.cat-holiday{position:relative;flex:1;min-height:92px;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;border-radius:7px}.grid .holiday-cell .chip.cat-holiday .top{position:absolute;top:7px;left:8px;right:8px;width:auto}.grid .holiday-cell .chip.cat-holiday .ttl{font-size:14px;margin:0}.grid .holiday-cell .chip.cat-holiday .det{display:none}
@media (max-width:820px){.floating-mode-switch{left:14px;right:auto;bottom:14px;grid-template-columns:repeat(3,42px)}.mode-option,.today-option{width:42px;height:42px}.mode-main{font-size:10px}.today-option .mode-main{font-size:7px}}
@media (min-width:821px) and (max-width:1400px){.floating-mode-switch{left:32px;grid-template-columns:repeat(3,36px)}.mode-option,.today-option{width:36px;height:34px}.mode-main{font-size:9px}.mode-sub{font-size:5.5px}.today-option .mode-main{font-size:7px}}
@media (orientation:landscape) and (max-height:540px){.day-hours{min-width:20px;padding:0 3px;font-size:6.3px;border-radius:6px}.dnum .day-hours:not([hidden])+.dweekday{margin-left:3px}.chip.layer-class.confirmed,.chip.layer-class.unconfirmed,.chip.layer-class.note{box-shadow:inset 4px 0 0 var(--context-color),0 1px 1px rgba(20,30,50,.04)}.floating-mode-switch{left:clamp(42px,3vw,112px);bottom:8px;grid-template-columns:repeat(3,34px);gap:2px;padding:3px}.mode-option,.today-option{width:34px;height:32px;border-radius:6px}.mode-main{font-size:8px}.today-option .mode-main{font-size:6px}.mode-sub{font-size:5px}.grid .holiday-cell .chip.cat-holiday{min-height:58px}.grid .holiday-cell .chip.cat-holiday .ttl{display:block;font-size:7px}}
@media print{.floating-mode-switch{display:none}}
.filter-jump-target{scroll-margin-top:12px}.top-option,.version-option{width:48px;height:44px;border:0;border-radius:8px;background:#fff;color:#0f7074;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:1px;cursor:pointer;font-family:inherit}.top-option:hover,.version-option:hover{background:#eef7f7}.top-option:focus-visible,.version-option:focus-visible{outline:3px solid #ffc857;outline-offset:1px}.top-option .mode-main{font-size:18px;line-height:.85}.version-option .mode-main{font-size:12px;line-height:1;font-weight:950}.top-option .mode-sub{font-size:7px;font-weight:850;line-height:1}@media (max-width:820px){.top-option,.version-option{width:42px;height:42px}.version-option .mode-main{font-size:11px}}@media (min-width:821px) and (max-width:1400px){.top-option,.version-option{width:36px;height:34px}.top-option .mode-main{font-size:13px}.version-option .mode-main{font-size:10px}.top-option .mode-sub{font-size:5.5px}}@media (orientation:landscape) and (max-height:540px){.top-option,.version-option{width:34px;height:32px;border-radius:6px}.top-option .mode-main{font-size:12px}.version-option .mode-main{font-size:9px}.top-option .mode-sub{font-size:5px}}
'''

CSS += r'''
.chip.changed-in-version{position:relative;outline:3px solid #f2a900;outline-offset:1px}.change-badge{position:absolute;z-index:3;top:4px;left:5px;padding:1px 4px;border:1px solid #8a5200;border-radius:3px;background:#ffd84d;color:#312300;font-size:7.5px;font-weight:950;line-height:1.15;white-space:nowrap;box-shadow:0 1px 2px rgba(35,27,0,.22)}.chip.erb-compact.changed-in-version .class-id{max-width:calc(100% - 76px)}.sample.changed-sample{border:3px solid #f2a900;background:#ffd84d}.filter.change-filter{border-color:#a96700;background:#fff4bd;color:#684000}.filter.change-filter.active{border-color:#6c4200;background:#f2a900;color:#241800}.pill.changed-pill{background:#ffd84d;color:#4d3200}.comparison-old{margin-top:16px;padding:12px;border:2px solid #e3aa24;border-radius:7px;background:#fff8d8;color:#3a3428;white-space:normal}.comparison-old strong{display:block;margin-bottom:5px;color:#6b4500}.comparison-old .old-status{margin-top:7px;color:#75694e;font-size:12px}.comparison-new{font-weight:800;color:#8a5300}.modal-current-label{display:block;margin-bottom:5px;color:#566273;font-size:11px;font-weight:850;text-transform:uppercase}.modal-current{white-space:pre-wrap}
@media (orientation:landscape) and (max-height:540px){.chip.changed-in-version{outline-width:1.5px}.change-badge{top:2px;left:2px;padding:0 2px;border-width:1px;font-size:4.7px}.chip.erb-compact.changed-in-version .class-id{max-width:calc(100% - 38px)}}
@media print{.change-badge{background:#ffd84d!important}.chip.changed-in-version{outline-color:#a96700}}
'''

CSS += r'''
@media (orientation:landscape) and (max-height:700px) and (max-width:1400px){
  html,body{-webkit-text-size-adjust:100%!important;text-size-adjust:100%!important}
  .wrap{padding:6px 5px 48px}.hero,.sub,.stats,.legend,.filters,.foot,.layer-controls,.course-code-legend{display:none}
  .month{padding:8px;margin-top:8px}.month h2{font-size:15px;margin:0 0 6px}
  .gridwrap{overflow:visible}.grid{min-width:0;grid-template-columns:repeat(7,minmax(0,1fr));gap:3px}
  .dow{font-size:8.5px;letter-spacing:0;padding:0}.cell{min-height:98px;height:auto;padding:2px;border-radius:5px;overflow:visible}
  .dnum{font-size:8.5px;gap:2px}.dnum .dmon{font-size:9px}.dnum .dday{font-size:8.5px}.dnum .dweekday{font-size:6.8px}
  .chip{padding:2px 3px 3px;border-radius:4px;overflow:visible}.chip.confirmed,.chip.unconfirmed{border-width:1.8px!important}.chip.note{border-width:1.3px!important}
  .chip .top{margin-bottom:1px}.chip .cat,.chip .status{font-size:5.8px;font-weight:550;letter-spacing:0}
  .class-id{margin-top:1px;padding:1px 2px;gap:2px;font-size:5.8px;border-radius:2px}.class-dot{width:4px;height:4px;flex-basis:4px}
  .chip .ttl,.chip .det{display:none}.chip .fulltxt{display:block;font-size:6.6px;font-weight:400;line-height:1.14;white-space:normal;overflow:visible;word-break:break-word;overflow-wrap:anywhere}
  .chip.erb-compact{gap:1px;padding:3px 4px 4px}.chip.erb-compact .class-id{min-height:10px;max-width:calc(100% - 14px);padding:1px 3px;font-size:6px;border-width:1.3px;border-radius:2px}
  .chip.erb-compact .class-dot{width:4px;height:4px;flex-basis:4px}.erb-meta,.erb-course,.erb-foot{font-size:6px;line-height:1.08}.chip.erb-compact .status{top:2px;right:3px;font-size:5.8px}
  .chip.layer-class,.chip.erb-compact.layer-class{padding-left:7px}.chip.layer-class.confirmed,.chip.layer-class.unconfirmed,.chip.layer-class.note{box-shadow:inset 4px 0 0 var(--context-color),0 1px 1px rgba(20,30,50,.04)}
  .overlap-group{gap:2px}.overlap-group.overlap-active{padding:1px 6px 1px 0}.overlap-group.overlap-active::after{width:4px;border-width:1.4px}
  .floating-mode-switch{left:clamp(42px,3vw,112px);bottom:8px;grid-template-columns:repeat(3,34px);gap:2px;padding:3px}
  .mode-option,.today-option,.top-option,.version-option{width:34px;height:32px;border-radius:6px}.mode-main{font-size:8px}.today-option .mode-main{font-size:6px}.mode-sub{font-size:5px}
  .top-option .mode-main,.version-option .mode-main{font-size:12px}.top-option .mode-sub,.version-option .mode-sub{font-size:5px}
  .grid .holiday-cell .chip.cat-holiday{min-height:58px}.grid .holiday-cell .chip.cat-holiday .ttl{display:block;font-size:7px}
  .chip.changed-in-version{outline-width:1.5px}.change-badge{top:2px;left:2px;padding:0 2px;border-width:1px;font-size:4.7px}.chip.erb-compact.changed-in-version .class-id{max-width:calc(100% - 38px)}
}
'''

CSS += r'''
.view-tabs{position:sticky;top:0;z-index:35;display:inline-grid;grid-template-columns:repeat(2,minmax(126px,1fr));gap:3px;margin:18px 0 2px;padding:4px;border:1px solid #cfd8e5;border-radius:8px;background:rgba(229,234,241,.96);box-shadow:0 2px 8px rgba(20,30,50,.10);backdrop-filter:blur(8px)}
.view-tab{min-height:38px;border:0;border-radius:6px;padding:7px 14px;background:transparent;color:#4c5a6b;font:inherit;font-size:13px;font-weight:850;cursor:pointer}.view-tab:hover{background:#f4f8fa}.view-tab:focus-visible{outline:3px solid #ffc857;outline-offset:1px}.view-tab.active{background:#fff;color:#0f6868;box-shadow:0 1px 3px rgba(20,30,50,.18)}.view-panel[hidden]{display:none!important}
.span-shell{margin-top:16px}.span-head{display:flex;align-items:flex-start;justify-content:space-between;gap:18px;margin-bottom:12px}.span-head h2{margin:0;font-size:20px;font-weight:850}.span-head p{max-width:780px;margin:4px 0 0;color:#667387;font-size:13px}.span-tools{display:flex;flex-direction:column;align-items:flex-end;gap:8px}.span-mode-switch{display:inline-grid;grid-template-columns:repeat(3,minmax(76px,1fr));gap:3px;padding:3px;border:1px solid #cfd8e5;border-radius:8px;background:#e5eaf1}.span-mode-option{min-height:34px;border:0;border-radius:6px;padding:5px 8px;background:transparent;color:#4c5a6b;font:inherit;font-size:10px;font-weight:900;line-height:1.05;cursor:pointer}.span-mode-option:hover{background:#f4f8fa}.span-mode-option:focus-visible{outline:3px solid #ffc857;outline-offset:1px}.span-mode-option.active{background:#0f7074;color:#fff;box-shadow:0 1px 3px rgba(20,30,50,.18)}.span-legend{display:flex;flex-wrap:wrap;justify-content:flex-end;gap:8px}.span-legend-item{display:flex;align-items:center;gap:7px;padding:6px 9px;border:1px solid #dce3ec;border-radius:6px;background:#fff;color:#566273;font-size:11.5px;font-weight:750;white-space:nowrap}.span-key{width:9px;height:22px;border-radius:999px;background:#244b50}.span-key.other{height:16px;background:#b9c4ce}.span-scroll{overflow-x:auto;overflow-y:visible;border:1px solid #dbe2ec;border-radius:8px;background:#fff;box-shadow:0 1px 3px rgba(20,30,50,.06);overscroll-behavior-inline:contain;scrollbar-gutter:stable}.span-table{min-width:2050px}.span-axis,.span-row{display:grid;grid-template-columns:280px minmax(1760px,1fr)}.span-axis{min-height:48px;border-bottom:1px solid #dbe2ec;background:#f7f9fc}.span-axis-label,.span-label{position:sticky;left:0;z-index:4;border-right:1px solid #dbe2ec;background:#fff}.span-axis-label{display:flex;align-items:center;padding:9px 13px;background:#f7f9fc;color:#697586;font-size:11px;font-weight:850;text-transform:uppercase}.span-axis-track,.span-track{position:relative;min-width:0}.span-axis-track{height:48px}.span-month{position:absolute;inset-block:0 auto;display:flex;align-items:center;justify-content:center;border-left:1px solid #cad4df;color:#5f6b7b;font-size:11px;font-weight:850;text-transform:uppercase}.span-month:last-child{border-right:1px solid #cad4df}.span-row{min-height:74px;border-bottom:1px solid #e5eaf0}.span-row:last-child{border-bottom:0}.span-row[hidden],.span-marker[hidden]{display:none!important}.span-label{display:flex;flex-direction:column;justify-content:center;padding:8px 13px}.span-label strong{color:#1d2734;font-size:12px;line-height:1.2}.span-course-name{margin-top:2px;color:#435165;font-size:10.5px;font-weight:750;line-height:1.22}.span-label small{margin-top:3px;color:#738092;font-size:10px;line-height:1.25}.span-track{height:74px;background-image:linear-gradient(to right,rgba(225,231,238,.75) 1px,transparent 1px);background-size:calc(100% / 8) 100%}.span-bar{position:absolute;top:19px;height:36px;min-width:28px;border:2px solid hsl(var(--span-hue),55%,38%);border-radius:999px;background:hsl(var(--span-hue),64%,91%);box-shadow:inset 0 1px 1px rgba(255,255,255,.8),0 1px 2px rgba(25,36,52,.14)}.span-bar::before,.span-bar::after{position:absolute;top:50%;z-index:1;transform:translateY(-50%);padding:1px 4px;border-radius:3px;background:rgba(255,255,255,.88);color:#475365;font-size:8px;font-weight:850;line-height:1.2;white-space:nowrap;pointer-events:none}.span-bar::before{content:attr(data-first-label);left:5px}.span-bar::after{content:attr(data-last-label);right:5px}.span-marker{position:absolute;top:50%;z-index:2;width:7px;height:17px;transform:translate(-50%,-50%);border:1px solid rgba(43,54,68,.35);border-radius:999px;background:#b9c4ce;box-shadow:0 0 0 1px rgba(255,255,255,.72);cursor:help}.span-marker.mine{z-index:3;width:9px;height:28px;border-color:#193e43;background:#244b50;box-shadow:0 0 0 2px rgba(255,255,255,.82)}.span-marker.unconfirmed{border-style:dashed}.span-marker:focus-visible{outline:3px solid #ffc857;outline-offset:2px}.span-empty{padding:24px;color:#667387;text-align:center}.span-view-active .floating-mode-switch,.span-view-active .actions{display:none}
@media (max-width:820px){.view-tabs{display:grid;width:100%;grid-template-columns:repeat(2,minmax(0,1fr));margin-top:14px}.view-tab{padding:7px 6px}.span-head{display:block}.span-tools{align-items:stretch;margin-top:10px}.span-mode-switch{width:100%}.span-legend{justify-content:flex-start}.span-scroll{width:100%}.span-table{min-width:1880px}.span-axis,.span-row{grid-template-columns:230px minmax(1640px,1fr)}.span-axis-label,.span-label{padding-left:10px;padding-right:10px}.span-label strong{font-size:11px}.span-course-name{font-size:9.5px}}
@media (orientation:landscape) and (max-height:700px) and (max-width:1400px) and (pointer:coarse){.view-tabs{position:sticky;display:inline-grid;width:auto;margin:3px 0;padding:2px}.view-tab{min-height:28px;padding:4px 9px;font-size:10px}.span-shell{margin-top:5px}.span-head h2{font-size:15px}.span-head p{font-size:10px}.span-tools{gap:4px}.span-mode-switch{grid-template-columns:repeat(3,minmax(60px,1fr));padding:2px}.span-mode-option{min-height:28px;padding:3px 5px;font-size:8px}.span-legend-item{padding:3px 6px;font-size:8px}.span-key{height:16px}.span-key.other{height:11px}.span-table{min-width:1760px}.span-axis,.span-row{grid-template-columns:205px minmax(1540px,1fr)}.span-row{min-height:58px}.span-track{height:58px}.span-bar{top:13px;height:32px}.span-label strong{font-size:9px}.span-course-name{font-size:7.5px}.span-label small{font-size:7px}}
@media (pointer:fine){.span-table{min-width:2050px}.span-axis,.span-row{grid-template-columns:280px minmax(1760px,1fr)}.span-row{min-height:74px}.span-track{height:74px}.span-bar{top:19px;height:36px}.span-label strong{font-size:12px}.span-course-name{font-size:10.5px}.span-label small{font-size:10px}}
@media print{.view-tabs{display:none}.view-panel[hidden]{display:block!important}.span-scroll{overflow:visible}.span-table{min-width:0}.span-axis,.span-row{grid-template-columns:180px minmax(0,1fr)}}
'''

CSS += r'''
.span-track{background-image:none;background-size:auto}.span-guide{position:absolute;inset-block:0;border-left:1px solid rgba(216,224,233,.82);pointer-events:none}.span-marker{padding:0;cursor:pointer}
'''

CSS += r'''
html{overflow-x:auto;overflow-y:scroll}
.time-slot{display:contents}.time-slot.slot-hidden{display:none}.transit-notice{margin:10px 0 12px;padding:8px 10px;border:1px solid #9fc7c6;border-left:5px solid #0f7074;border-radius:5px;background:#f4fbfa;color:#36545a;font-size:11.5px;font-weight:750}.transit-bar{margin:3px 0;padding:3px 6px;border:1px solid #91bab8;border-radius:3px;background:#edf8f7;color:#255b5d;font-size:8.5px;font-weight:850;line-height:1.2;text-align:center}.transit-bar.tight{border-color:#c9342d;background:#fff0ee;color:#951f1a}.transit-bar strong{font-weight:950}
.filter-heading{display:flex;align-items:end;justify-content:space-between;gap:12px;margin-top:22px}.filter-heading .section-h{margin:0}.filter-master{display:flex;align-items:center;gap:8px;flex-wrap:wrap}.filter-groups{display:grid;gap:7px;margin:10px 0 16px}.filter-group{display:grid;grid-template-columns:92px minmax(0,1fr);align-items:start;gap:10px;padding-top:7px;border-top:1px solid #dbe2ec}.filter-group:first-child{border-top:0}.filter-group-label{padding:8px 0;color:#657285;font-size:11px;font-weight:900;text-transform:uppercase}.filter-group-buttons{display:flex;gap:7px;flex-wrap:wrap}.filter-group .filter{padding:6px 10px}
.upcoming-summary{margin:18px 0 4px}.upcoming-summary .section-h{margin:0 2px 9px}.upcoming-list{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:7px}.upcoming-course{display:grid;grid-template-columns:58px minmax(0,1fr) 40px minmax(112px,.72fr);align-items:center;gap:8px;min-width:0;border:2.5px solid #1d2734;border-radius:6px;padding:7px 9px;background:#fff;color:#263343;text-align:left;font:inherit;cursor:pointer;box-shadow:0 1px 2px rgba(20,30,50,.05)}.upcoming-course.unconfirmed,.upcoming-course.mixed{border-style:dashed}.upcoming-course:hover{background:#f7fbfb}.upcoming-course:focus-visible{outline:3px solid #ffc857;outline-offset:2px}.upcoming-course.active{background:#0f7074;color:#fff}.upcoming-course.active.unconfirmed,.upcoming-course.active.mixed{border-width:3px;border-style:dashed;border-color:#fff;box-shadow:0 0 0 3px #e5a900,0 1px 3px rgba(20,30,50,.24)}.upcoming-date{font-size:13px;font-weight:950;color:#0f7074;white-space:nowrap}.upcoming-course.active .upcoming-date{color:#fff}.upcoming-course-copy,.upcoming-place{display:flex;min-width:0;flex-direction:column}.upcoming-course-copy strong{font-size:12px;font-weight:950;line-height:1.15}.upcoming-course-copy span{margin-top:2px;color:#536174;font-size:9.5px;font-weight:750;line-height:1.18;overflow-wrap:anywhere}.upcoming-place{padding-left:8px;border-left:1px solid #dce3ec}.upcoming-place strong{font-size:9.5px;font-weight:900;line-height:1.16}.upcoming-place span{margin-top:2px;color:#647285;font-size:9px;font-weight:700;line-height:1.16;overflow-wrap:anywhere}.upcoming-course.active .upcoming-course-copy span,.upcoming-course.active .upcoming-place span{color:#e4f3f3}.upcoming-course.active .upcoming-place{border-left-color:#7eb1b2}.upcoming-language{justify-self:center;border:1px solid #a8b5c4;border-radius:4px;padding:2px 4px;background:#eef3f7;color:#405064;font-size:8.5px;font-weight:950}.upcoming-course.active .upcoming-language{border-color:#b9d7d8;background:#fff;color:#0f7074}
.span-shell{--span-label-width:280px;--span-timeline-width:1760px;margin-top:16px}.span-head{display:block;max-width:980px;margin-bottom:10px}.span-head p{max-width:900px}.span-control-bar{position:sticky;left:16px;z-index:8;display:flex;align-items:flex-start;gap:8px;max-width:calc(100vw - 52px);margin-bottom:10px;flex-wrap:wrap}.span-tools{display:flex;flex-direction:row;align-items:center;justify-content:flex-start;gap:8px;flex-wrap:wrap}.span-tool-button,.span-zoom button,.span-course-picker summary,.span-course-actions button{min-height:40px;border:1px solid #cfd8e5;border-radius:6px;background:#fff;color:#405064;font:inherit;font-size:11px;font-weight:850;cursor:pointer}.span-tool-button{padding:7px 10px}.span-tool-button[aria-pressed="false"]{background:#405064;color:#fff}.span-zoom{display:grid;grid-template-columns:40px 64px 40px;gap:3px}.span-zoom button{padding:0;font-size:19px}.span-zoom .span-zoom-value{font-size:10px}.span-tool-button:focus-visible,.span-zoom button:focus-visible,.span-course-picker summary:focus-visible,.span-course-actions button:focus-visible{outline:3px solid #ffc857;outline-offset:1px}.span-course-picker{position:relative}.span-course-picker summary{display:flex;align-items:center;gap:7px;padding:7px 10px;list-style:none}.span-course-picker summary::-webkit-details-marker{display:none}.span-course-picker summary::before{content:"+";font-size:15px}.span-course-picker[open] summary::before{content:"−"}.span-course-picker-body{position:absolute;top:46px;left:0;z-index:20;width:min(680px,calc(100vw - 48px));padding:10px;border:1px solid #cfd8e5;border-radius:7px;background:#fff;box-shadow:0 8px 24px rgba(20,30,50,.20)}.span-course-actions{display:flex;gap:6px;margin-bottom:8px}.span-course-actions button{min-height:32px;padding:4px 9px}.span-course-options{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));max-height:min(52vh,440px);overflow:auto;border-top:1px solid #e2e7ef}.span-course-toggle{display:flex;align-items:flex-start;gap:8px;min-width:0;padding:8px;border-bottom:1px solid #e8edf3;cursor:pointer}.span-course-toggle:nth-child(odd){border-right:1px solid #e8edf3}.span-course-toggle input{width:17px;height:17px;margin:1px 0 0;accent-color:#0f7074;flex:0 0 auto}.span-course-toggle span{display:flex;min-width:0;flex-direction:column}.span-course-toggle strong{font-size:10.5px;line-height:1.2}.span-course-toggle small{margin-top:2px;color:#748092;font-size:9px}.span-legend{justify-content:flex-start}.span-scroll{width:max-content;min-width:100%;overflow:visible;border:1px solid #dbe2ec;border-radius:8px;scrollbar-gutter:auto}.span-table{width:calc(var(--span-label-width) + var(--span-timeline-width));min-width:calc(var(--span-label-width) + var(--span-timeline-width))}.span-axis,.span-row{grid-template-columns:var(--span-label-width) var(--span-timeline-width)}.span-axis{border-bottom:1px solid #dbe2ec}.span-row{border-bottom:1px solid #dbe2ec}.span-axis-label,.span-label{border-right:1px solid #dbe2ec}.span-month{border-left:1px solid #dbe2ec}.span-month:first-child{border-left:0}.span-month:last-child{border-right:0}.span-guide{border-left:1px solid #dbe2ec}.span-shell.hide-span-labels{--span-label-width:0px}.span-shell.hide-span-labels .span-axis-label,.span-shell.hide-span-labels .span-label{display:none}.span-row[data-course-hidden="1"]{display:none!important}
@media (pointer:fine) and (min-width:821px){body.span-view-active .wrap{overflow:visible}.span-view-active .span-scroll{width:max-content;min-width:100%;overflow:visible}.span-view-active .span-control-bar{position:sticky;left:16px}}
@media (max-width:1100px){.upcoming-list{grid-template-columns:1fr}}
@media (max-width:820px){html{overflow-x:hidden}.transit-notice{font-size:10.5px}.transit-bar{font-size:8px}.filter-heading{align-items:flex-start}.filter-groups{gap:5px}.filter-group{grid-template-columns:62px minmax(0,1fr);gap:6px}.filter-group-label{font-size:10px}.upcoming-course{grid-template-columns:54px minmax(0,1fr) 38px;padding:7px}.upcoming-place{grid-column:2/-1;padding:5px 0 0;border-top:1px solid #dce3ec;border-left:0}.upcoming-course.active .upcoming-place{border-top-color:#7eb1b2;border-left:0}.span-shell{--span-label-width:230px;--span-timeline-width:1640px}.span-control-bar{position:static;display:block;max-width:none}.span-tools{align-items:stretch}.span-mode-switch{width:100%}.span-zoom{width:100%;grid-template-columns:44px minmax(72px,1fr) 44px}.span-tool-button,.span-course-picker{width:100%}.span-course-picker summary{justify-content:center}.span-course-picker-body{position:fixed;top:76px;left:12px;right:12px;width:auto;max-height:calc(100vh - 94px);overflow:auto}.span-course-options{grid-template-columns:1fr;max-height:none}.span-course-toggle:nth-child(odd){border-right:0}.span-scroll{width:100%;min-width:0;overflow-x:auto;overflow-y:visible;overscroll-behavior-inline:contain;scrollbar-gutter:stable}.span-table{width:calc(var(--span-label-width) + var(--span-timeline-width));min-width:calc(var(--span-label-width) + var(--span-timeline-width))}.span-axis,.span-row{grid-template-columns:var(--span-label-width) var(--span-timeline-width)}}
@media (orientation:landscape) and (max-height:700px) and (max-width:1400px) and (pointer:coarse){.upcoming-summary,.filter-heading,.filter-groups{display:none}.span-shell{--span-label-width:205px;--span-timeline-width:1540px}.span-control-bar{display:flex}.span-tools{gap:4px}.span-tool-button,.span-zoom button,.span-course-picker summary{min-height:30px;font-size:8px}.span-zoom{width:auto;grid-template-columns:30px 48px 30px}.span-course-picker{width:auto}.span-course-picker-body{top:44px}.span-course-options{grid-template-columns:repeat(2,minmax(0,1fr))}.span-scroll{width:100%;min-width:0;overflow-x:auto;overflow-y:visible;overscroll-behavior-inline:contain;scrollbar-gutter:stable}}
@media print{html{overflow:visible}.upcoming-summary,.filter-heading,.filter-groups,.span-control-bar{display:none}.span-scroll{width:100%;overflow:visible}.span-table{width:100%;min-width:0}.span-axis,.span-row{grid-template-columns:180px minmax(0,1fr)}}
'''

CSS += r'''
.course-code-heading{display:flex;align-items:center;gap:8px;margin-top:22px}.course-code-heading .section-h{margin:0}.course-code-count{border:1px solid #b8c5d3;border-radius:4px;padding:2px 6px;background:#fff;color:#526174;font-size:9px;font-weight:900}
.class-summary{margin:18px 0 4px}.class-summary-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin:0 2px 9px}.class-summary-head .section-h{margin:0}.class-summary-key{display:flex;align-items:center;gap:8px;flex-wrap:wrap}.class-summary-key span{display:inline-flex;align-items:center;gap:5px;color:#526174;font-size:9px;font-weight:850}.class-summary-key i{display:inline-block;width:20px;height:10px;border:2px solid #0f7074;border-radius:2px;background:#e9f7f5}.class-summary-key .pending i{border:2px dashed #a64b00;background:#fff0d2}
.class-summary-list{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:8px}.class-summary-card{position:relative;display:grid;grid-template-columns:62px minmax(0,1fr) 42px 38px minmax(112px,.72fr) auto;align-items:center;gap:8px;min-width:0;border:2px solid #0f7074;border-radius:5px;padding:8px 9px;background:#edf9f7;color:#263343;text-align:left;font:inherit;cursor:pointer;box-shadow:0 1px 2px rgba(20,30,50,.06)}.class-summary-card.confirmed{border-style:solid;border-color:#0f7074;background:#eaf7f5}.class-summary-card.unconfirmed{border-width:3px;border-style:dashed;border-color:#a64b00;background:#fff3da}.class-summary-card.mixed{border-width:3px;border-style:dashed;border-color:#6b4bb5;background:#f3efff}.class-summary-card.completed{border-color:#64748b;background:#f3f6f9}.class-summary-card:hover{filter:brightness(.985)}.class-summary-card:focus-visible{outline:3px solid #ffc857;outline-offset:2px}.class-summary-card.active,.class-summary-card.active.confirmed,.class-summary-card.active.unconfirmed,.class-summary-card.active.mixed,.class-summary-card.active.completed{background:inherit;color:inherit;outline:4px solid #ffc857;outline-offset:1px;box-shadow:0 0 0 1px #1d2734,0 2px 5px rgba(20,30,50,.2)}.class-summary-card.active .upcoming-date{color:#0f7074}.class-summary-card.active .upcoming-course-copy span,.class-summary-card.active .upcoming-place span{color:#536174}.class-summary-card.active .upcoming-place{border-left-color:#b9c5d1}.class-summary-card.active .upcoming-language{border-color:#a8b5c4;background:#eef3f7;color:#405064}.provider-badge{display:flex;width:36px;height:30px;align-items:center;justify-content:center;border:2px solid #1d2734;border-radius:5px;color:#fff;font-size:12px;font-weight:950;line-height:1;box-shadow:0 1px 2px rgba(20,30,50,.18)}.provider-badge.provider-ca{background:#006f78}.provider-badge.provider-mc{background:#6552a3}.summary-status{justify-self:end;border-radius:4px;padding:3px 6px;color:#fff;font-size:8px;font-weight:950;letter-spacing:0;white-space:nowrap}.confirmed .summary-status{background:#0f7074}.unconfirmed .summary-status{background:#a64b00}.mixed .summary-status{background:#6b4bb5}.completed .summary-status{background:#64748b}.completed-summary{margin-top:20px}.completed-summary .class-summary-card{opacity:.94}.completed-summary .upcoming-date{color:#526174}
.span-shell{--span-day-width:16px;--span-timeline-width:3920px}.span-control-bar{align-items:center}.span-month-controls{display:flex;align-items:center;gap:3px;flex-wrap:wrap}.span-month-controls-label{margin-right:3px;color:#5f6d7f;font-size:10px;font-weight:900}.span-month-toggle{min-width:40px;min-height:34px;border:1px solid #aebdcc;border-radius:5px;padding:4px 7px;background:#0f7074;color:#fff;font:inherit;font-size:9px;font-weight:900;cursor:pointer}.span-month-toggle[aria-pressed="false"]{background:#fff;color:#687689;text-decoration:line-through}.span-month-toggle:focus-visible{outline:3px solid #ffc857;outline-offset:1px}.span-axis{min-height:58px}.span-axis-track{display:flex;height:58px;width:var(--span-timeline-width);overflow:hidden}.span-month{position:relative;inset:auto;display:block;flex:0 0 calc(var(--month-days) * var(--span-day-width));width:calc(var(--month-days) * var(--span-day-width));height:58px;border-left:1px solid #aebdcc;color:#4f5d70;text-align:center}.span-month[hidden],.span-track-month[hidden]{display:none!important}.span-month>strong{display:block;height:24px;padding-top:5px;border-bottom:1px solid #cfd8e5;font-size:10px;font-weight:950}.span-days{display:grid;grid-template-columns:repeat(var(--month-days),var(--span-day-width));height:34px}.span-day{display:flex;align-items:center;justify-content:center;border-left:1px solid #e0e6ed;color:#708094;font-size:7px;font-weight:750;overflow:hidden}.span-day:first-child{border-left:0}.span-row{min-height:76px}.span-label{position:sticky;padding-right:38px}.span-row-toggle{position:absolute;top:8px;right:8px;width:24px;height:24px;border:1px solid #b7c3d0;border-radius:4px;background:#fff;color:#4c5b6d;font:inherit;font-size:13px;font-weight:900;line-height:1;cursor:pointer}.span-row-toggle:hover{background:#fff0ee;color:#a1261f}.span-row-toggle:focus-visible{outline:3px solid #ffc857;outline-offset:1px}.span-track{height:76px;width:var(--span-timeline-width);background:none}.span-track-grid{position:absolute;inset:0;display:flex;z-index:0;overflow:hidden;pointer-events:none}.span-track-month{display:block;flex:0 0 calc(var(--month-days) * var(--span-day-width));width:calc(var(--month-days) * var(--span-day-width));height:100%;border-left:1px solid #aebdcc;background-image:repeating-linear-gradient(to right,transparent 0,transparent calc(var(--span-day-width) - 1px),#e0e6ed calc(var(--span-day-width) - 1px),#e0e6ed var(--span-day-width))}.span-track-month:first-child{border-left:0}.span-bar{top:20px;z-index:1;height:36px;min-width:var(--span-day-width);border-radius:4px}.span-marker{width:8px;height:18px;border-radius:3px}.span-marker.mine{width:10px;height:28px;border-radius:3px}.span-bar::before,.span-bar::after{font-size:7px}.span-table{width:calc(var(--span-label-width) + var(--span-timeline-width));min-width:calc(var(--span-label-width) + var(--span-timeline-width))}.span-axis,.span-row{grid-template-columns:var(--span-label-width) var(--span-timeline-width)}
.completed-summary .class-summary-card{grid-template-columns:92px minmax(0,1fr) 42px 38px minmax(112px,.72fr) auto}
@media(max-width:1100px){.class-summary-list{grid-template-columns:1fr}}
@media(max-width:820px){.course-code-heading{align-items:flex-start}.class-summary-head{align-items:flex-start;flex-direction:column}.class-summary-card{grid-template-columns:56px minmax(0,1fr) 38px 38px auto;padding:8px}.class-summary-card .upcoming-place{grid-column:2/-1}.summary-status{grid-column:5}.span-month-controls{width:100%}.span-month-controls-label{width:100%}.span-month-toggle{flex:1;min-width:44px}.span-row-toggle{top:6px;right:6px}.span-shell{--span-day-width:16px}}
@media(max-width:820px){.completed-summary .class-summary-card{grid-template-columns:82px minmax(0,1fr) 38px 38px auto}}
@media (orientation:landscape) and (max-height:700px) and (max-width:1400px) and (pointer:coarse){.class-summary{display:none}.span-month-toggle{min-height:28px;font-size:8px}.span-axis{min-height:50px}.span-axis-track,.span-month{height:50px}.span-month>strong{height:21px;padding-top:4px}.span-days{height:29px}.span-row{min-height:60px}.span-track{height:60px}.span-bar{top:13px;height:34px}}
@media print{.class-summary{display:none}.span-month-controls,.span-row-toggle{display:none}}
'''

CSS += r'''
.span-month-controls{order:10}.span-month-toggle{position:relative;display:inline-block;min-width:44px;min-height:34px;border:0;padding:0;background:transparent;color:inherit}.span-month-toggle input{position:absolute;width:1px;height:1px;opacity:0;pointer-events:none}.span-month-toggle span{display:flex;min-width:44px;min-height:34px;align-items:center;justify-content:center;border:1px solid #aebdcc;border-radius:5px;background:#fff;color:#687689;font-size:9px;font-weight:900;text-decoration:line-through;cursor:pointer}.span-month-toggle input:checked+span{border-color:#0f7074;background:#0f7074;color:#fff;text-decoration:none}.span-month-toggle input:focus-visible+span{outline:3px solid #ffc857;outline-offset:1px}
.span-course-picker{position:static;order:20;flex:1 0 100%;width:100%;margin-top:2px;border:1px solid #cfd8e5;border-radius:7px;background:#fff;overflow:hidden;box-shadow:0 1px 3px rgba(20,30,50,.08)}.span-course-picker-head{display:flex;align-items:center;gap:10px;min-height:42px;padding:7px 10px;border-bottom:1px solid #dce3ec;background:#f7f9fc}.span-course-picker-head>strong{font-size:11px;font-weight:950}.span-course-picker-head>#spanCourseCount{color:#0f7074;font-size:10px;font-weight:900}.span-course-picker-head .span-course-actions{margin:0 0 0 auto}.span-course-picker-body{position:static;width:auto;padding:0;border:0;border-radius:0;background:#fff;box-shadow:none}.span-course-options{display:grid;grid-template-columns:repeat(4,minmax(180px,1fr));max-height:260px;overflow:auto;border-top:0}.span-course-toggle{min-height:52px;border-right:1px solid #e8edf3;border-bottom:1px solid #e8edf3;background:#f4f6f8}.span-course-toggle:has(input:checked){border-left:4px solid #0f7074;background:#eaf7f5}.span-course-toggle input{accent-color:#0f7074}.span-course-toggle input+span strong::after{content:" OFF";margin-left:4px;color:#a1261f;font-size:8px}.span-course-toggle input:checked+span strong::after{content:" ON";color:#0f7074}.span-course-toggle input:not(:checked)+span{opacity:.68}.span-legend{order:15}.span-row-toggle{display:none!important}
@media(max-width:1100px){.span-course-options{grid-template-columns:repeat(3,minmax(180px,1fr))}}
@media(max-width:820px){.span-month-toggle{flex:1}.span-month-toggle span{width:100%}.span-course-options{grid-template-columns:1fr;max-height:300px}.span-course-picker-head{align-items:flex-start;flex-wrap:wrap}.span-course-picker-head .span-course-actions{width:100%;margin-left:0}.span-course-picker-head .span-course-actions button{flex:1}}
@media (orientation:landscape) and (max-height:700px) and (max-width:1400px) and (pointer:coarse){.span-course-picker{width:100%}.span-course-options{grid-template-columns:repeat(3,minmax(160px,1fr));max-height:180px}.span-course-toggle{min-height:40px;padding:5px}.span-month-toggle span{min-height:28px}}
@media print{.span-course-picker{display:none}}
'''

CSS += r'''
.span-course-breakdown{color:#647285;font-size:10px;font-weight:800}.span-course-toggle:has(input:checked){background:#f8fbfc}.span-course-toggle.status-upcoming{border-left:5px solid #0f7074;background:#e8f7f5}.span-course-toggle.status-pending{border-left:5px dashed #b65a00;background:#fff1d7}.span-course-toggle.status-completed{border-left:5px solid #8a96a5;background:#edf1f5}.span-course-toggle.status-context{border-left:5px solid #6f60aa;background:#f0eef9}.span-course-toggle.type-sen{box-shadow:inset 0 -4px 0 #2792a0}.span-status-tag{display:inline-flex;align-items:center;align-self:flex-start;margin-top:3px;padding:1px 5px;border-radius:3px;background:#fff;color:#526073;font-size:8px;font-weight:900;text-transform:uppercase}.span-bar-label{position:absolute;left:50%;top:50%;z-index:1;max-width:calc(100% - 56px);transform:translate(-50%,-50%);padding:2px 6px;border:1px solid hsla(var(--span-hue),55%,38%,.45);border-radius:4px;background:rgba(255,255,255,.90);color:hsl(var(--span-hue),55%,28%);font-size:8px;font-weight:950;line-height:1.15;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;pointer-events:none}
@media(max-width:820px){.span-course-breakdown{flex-basis:100%}.span-bar-label{font-size:7px}}
'''

CSS += r'''
.version-menu{width:100%;margin:14px 0 2px;border:1px solid #cfd8e5;border-radius:8px;background:#fff;box-shadow:0 1px 3px rgba(20,30,50,.08);overflow:hidden}.version-menu>summary{display:grid;grid-template-columns:auto auto minmax(0,1fr) auto;align-items:center;gap:10px;min-height:46px;padding:8px 12px;list-style:none;cursor:pointer}.version-menu>summary::-webkit-details-marker{display:none}.version-menu>summary:hover{background:#f7fbfb}.version-menu>summary:focus-visible{outline:3px solid #ffc857;outline-offset:-3px}.version-menu-kicker{color:#6c7888;font-size:10px;font-weight:900;text-transform:uppercase}.version-menu-current{border-radius:5px;background:#0f7074;color:#fff;padding:3px 7px;font-size:12px;font-weight:950}.version-menu-summary{min-width:0;color:#405064;font-size:12px;font-weight:750;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.version-menu-arrow{color:#0f7074;font-size:15px;font-weight:950;transition:transform .15s ease}.version-menu[open] .version-menu-arrow{transform:rotate(180deg)}.version-menu-list{max-height:min(52vh,460px);overflow:auto;border-top:1px solid #dce3ec}.version-menu-item{width:100%;display:grid;grid-template-columns:150px minmax(0,1fr) auto;align-items:center;gap:10px;min-height:48px;padding:7px 12px;border:0;border-bottom:1px solid #e7ebf1;background:#fff;color:#263343;text-align:left;font:inherit;cursor:pointer}.version-menu-item:last-child{border-bottom:0}.version-menu-item:hover{background:#f4f9f9}.version-menu-item:focus-visible{outline:3px solid #ffc857;outline-offset:-3px}.version-menu-item.current{background:#e9f6f5}.version-menu-item strong{font-size:12px;font-weight:900}.version-menu-item span{min-width:0;color:#637084;font-size:11px;line-height:1.25}.version-menu-item.current span{color:#3d6567}.version-menu-badge{border-radius:999px;background:#0f7074;color:#fff!important;padding:2px 7px;font-size:9px!important;font-weight:900;white-space:nowrap}
@media (max-width:820px){.version-menu{margin-top:10px}.version-menu>summary{grid-template-columns:auto auto minmax(0,1fr) auto;gap:7px;padding:7px 9px}.version-menu-kicker{display:none}.version-menu-summary{font-size:11px}.version-menu-item{grid-template-columns:118px minmax(0,1fr);gap:7px;padding:7px 9px}.version-menu-item .version-menu-badge{grid-column:1/-1;justify-self:start}}
@media (orientation:landscape) and (max-height:700px) and (max-width:1400px){.version-menu{margin:4px 0}.version-menu>summary{min-height:32px;padding:4px 7px}.version-menu-current{padding:2px 5px;font-size:9px}.version-menu-summary{font-size:9px}.version-menu-arrow{font-size:11px}.version-menu-list{max-height:55vh}.version-menu-item{min-height:36px;padding:4px 7px}.version-menu-item strong{font-size:9px}.version-menu-item span{font-size:8px}}
@media print{.version-menu{display:none}}
'''

TIME_RANGE_RE = re.compile(r"(?<!\d)(2[0-3]|[01]?\d):?([0-5]\d)\s*(am|pm)?\s*-\s*(2[0-3]|[01]?\d):?([0-5]\d)(?!\d)\s*(am|pm)?", re.I)
TEACHER_RE = re.compile(r"\b(Garett|Garrett|Andy|Calvin|Mike(?:\s+Sir)?)\b", re.I)
NOTE_WORD_RE = re.compile(
    r"test|exam|presentation|discussion|assessment|report|cancel|substitut|"
    r"\u6301\u7e8c|\u671f\u672b|\u7b46\u8a66|\u5be6\u52d9\u8a66|\u8a55\u4f30|"
    r"\u532f\u5831|\u5831\u544a|\u5c0f\u7d44|\u5c08\u984c",
    re.I,
)
CANCELLED_TEACHING_RE = re.compile(r"\b(?:cancelled|canceled)\b|取消", re.I)


def clean_location(title, category):
    if category == "ymca":
        return "YMCA"
    if category == "dgs":
        return "DGS"
    if category in {"holiday", "mike"}:
        return "-"
    value = str(title or "")
    code_m = COURSE_CODE_RE.search(value)
    if code_m:
        value = value[:code_m.start()]
    value = re.sub(r"\([^)]*\)|\[[^\]]*\]", "", value)
    value = re.sub(r"\s*-\s*L\s*\d+.*$", "", value, flags=re.I)
    value = re.sub(r"\s*-\s*", "-", value)
    return value.strip(" ,-/") or "-"


def display_clock(hour, minute, marker):
    hour = int(hour)
    minute = int(minute)
    marker = (marker or "").lower()
    if marker == "pm" and hour < 12:
        hour += 12
    elif marker == "am" and hour == 12:
        hour = 0
    return f"{hour:02d}:{minute:02d}"


def display_times(text):
    ranges = []
    for match in TIME_RANGE_RE.finditer(str(text or "")):
        start = display_clock(match.group(1), match.group(2), match.group(3))
        end = display_clock(match.group(4), match.group(5), match.group(6))
        ranges.append(f"{start}-{end}")
    return " / ".join(ranges) or "-"


def display_notes(text):
    notes = re.findall(r"\[([^\]]+)\]", str(text or ""))
    for value in re.findall(r"\(([^)]+)\)", str(text or "")):
        if NOTE_WORD_RE.search(value):
            notes.append(value)
    return list(dict.fromkeys(note.strip() for note in notes if note.strip()))


def event_fields(ev):
    text = str(ev.get("text") or "")
    title = str(ev.get("title") or "")
    category = ev.get("category")
    parts = slash_parts(text)

    teacher_m = TEACHER_RE.search(text)
    teacher = str(ev.get("teacher") or (teacher_m.group(1) if teacher_m else "-"))
    if teacher.lower() in {"garett", "garrett"}:
        teacher = "Garett"
    elif teacher.lower().startswith("mike"):
        teacher = "Mike Sir"

    if category == "ymca":
        course_name = "SEN"
    elif category == "dgs":
        course_name = re.sub(r"^DGS\s*", "", title, flags=re.I).strip() or "Unreal"
    elif category == "holiday":
        course_name = title or "Public Holiday"
    elif category == "mike":
        course_name = re.sub(r"^Mike\s+Sir\s*", "", title, flags=re.I)
        course_name = re.sub(r"\s*-\s*L\s*\d+.*$", "", course_name, flags=re.I).strip() or "AI Lesson"
    else:
        course_name = parts[1] if len(parts) > 1 else (ev.get("detail") or title)
        code_m = COURSE_CODE_RE.search(course_name)
        if code_m:
            course_name = course_name[:code_m.start()]
        time_m = TIME_RANGE_RE.search(course_name)
        if time_m:
            course_name = course_name[:time_m.start()]
        course_name = course_name.strip(" ,-/") or "-"

    class_label = ev.get("group_label") or "-"
    if category in {"holiday", "school", "mike"}:
        class_label = "-"
    lesson_m = LESSON_RE.search(text)
    if lesson_m:
        lesson = f"Lesson {lesson_m.group(1)}"
    elif re.search(r"\bLesson\s+TBC\b", text, re.I):
        lesson = "Lesson TBC"
    else:
        lesson = "Lesson -"
    location = clean_location(title, category)
    if category == "methodist":
        location = re.sub(r"^循道(?:衞理中心)?", "循道衞理中心", location)
    return {
        "class_label": class_label,
        "location": location,
        "teacher": teacher,
        "course_name": course_name,
        "time": display_times(text),
        "lesson": lesson,
        "notes": display_notes(text),
        "helper": str(ev.get("helper") or "").strip(),
    }


def event_layer(ev):
    if ev.get("layer") in {"mine", "class"}:
        return ev["layer"]
    if ev.get("category") == "mike":
        return "other"
    if ev.get("category") != "erb":
        return "mine"
    return "class" if event_fields(ev)["teacher"] not in {"Garett", "-"} else "mine"


def time_range_intervals(text):
    def minutes(hour, minute, marker):
        hour = int(hour)
        minute = int(minute)
        marker = (marker or "").lower()
        if marker == "pm" and hour < 12:
            hour += 12
        elif marker == "am" and hour == 12:
            hour = 0
        return hour * 60 + minute

    intervals = []
    seen = set()
    for match in TIME_RANGE_RE.finditer(str(text or "")):
        start = minutes(match.group(1), match.group(2), match.group(3))
        end = minutes(match.group(4), match.group(5), match.group(6))
        interval = (start, end)
        if end > start and interval not in seen:
            seen.add(interval)
            intervals.append(interval)
    return intervals


def event_interval(ev):
    intervals = time_range_intervals(ev.get("text"))
    if not intervals:
        return None
    return min(start for start, _end in intervals), max(end for _start, end in intervals)


def teaching_intervals(ev):
    text = str(ev.get("text") or "")
    if event_layer(ev) != "mine" or ev.get("category") in {"holiday", "mike"}:
        return []
    if CANCELLED_TEACHING_RE.search(text):
        return []
    return time_range_intervals(text)


def event_slot(ev):
    interval = event_interval(ev)
    if interval is None:
        return "other"
    start, _end = interval
    if start < 13 * 60:
        return "morning"
    if start < 18 * 60:
        return "afternoon"
    return "night"


def centre_code(ev):
    text = str(ev.get("text") or "")
    category = ev.get("category")
    if category == "ymca":
        return "ymca_yau_ma_tei"
    if "上水彩園" in text or "彩園邨彩湖樓" in text:
        return "sheung_shui"
    if "四海" in text:
        return "four_seas"
    if "彩雲" in text:
        return "choi_wan"
    if "灣仔" in text or "循道" in text:
        return "wan_chai"
    if "藍田" in text:
        return "lam_tin"
    if "順天" in text:
        return "shun_tin"
    return ""


def chip(ev):
    st = ev['status']
    mark = '✓' if st == 'confirmed' else '?' if st == 'unconfirmed' else '•'
    title_html = ev.get("title_html") or ehtml(ev["title"])
    detail_html = ev.get("detail_html") or ehtml(ev["detail"])
    full_html = ev.get("html") or ehtml(ev["text"])
    red_cls = " has-red" if ev.get("red") else ""
    layer = event_layer(ev)
    layer_cls = f" layer-{layer}"
    interval = event_interval(ev)
    lesson_intervals = teaching_intervals(ev)
    lesson_intervals_attr = ",".join(f"{start}-{end}" for start, end in lesson_intervals)
    interval_attrs = (
        f' data-start="{interval[0]}" data-end="{interval[1]}"'
        if interval else ' data-start="" data-end=""'
    )
    layer_attrs = (f' data-layer="{layer}" data-erb="{1 if ev["category"] == "erb" else 0}"'
                   f' data-course="{1 if ev["category"] in {"erb", "methodist"} else 0}"'
                   f' data-centre="{ehtml(centre_code(ev))}"{interval_attrs}'
                   f' data-teaching-intervals="{lesson_intervals_attr}"'
                   f' data-source="{ehtml(ev.get("source", ""))}"')
    changed = bool(ev.get("changed_in_version"))
    comparison_cls = " changed-in-version" if changed else ""
    comparison_attrs = (
        f' data-changed="{1 if changed else 0}"'
        f' data-change-kind="{ehtml(ev.get("change_kind", ""))}"'
        f' data-previous="{ehtml(ev.get("previous_text", ""))}"'
        f' data-previous-status="{ehtml(ev.get("previous_status", ""))}"'
    )
    comparison_badge = f'<span class="change-badge" title="Changed in {COMPARE_LABEL}">Δ {COMPARE_LABEL}</span>' if changed else ""
    if ev["category"] not in {"erb", "methodist"}:
        teacher_suffix = ""
        if layer == "class" and ev.get("teacher"):
            teacher_suffix = f' / <span class="context-teacher">Teacher: {ehtml(ev["teacher"])}</span>'
        return (f'<div class="chip {st} cat-{ev["category"]} grp-{ev["group"]}{red_cls}{layer_cls}{comparison_cls}" tabindex="0" role="button" '
                f'data-date="{ehtml(ev["date"])}" data-status="{ehtml(st)}" data-cat="{ehtml(ev["category_label"])}" data-group="{ehtml(ev["group"])}" data-group-label="{ehtml(ev["group_label"])}" data-text="{ehtml(ev["text"])}" data-html="{ehtml(full_html)}"{layer_attrs}{comparison_attrs}>'
                f'{comparison_badge}<div class="top"><span class="cat">{ehtml(ev["category_label"])}</span><span class="status">{mark}</span></div>'
                f'<div class="ttl">{title_html}</div><div class="det">{detail_html}{teacher_suffix}</div><div class="fulltxt">{full_html}</div></div>')

    fields = event_fields(ev)
    class_label = fields["class_label"]
    class_hue = zlib.crc32(class_label.encode("utf-8")) % 360
    identity_html = (f'<span class="class-id card-course-filter" role="button" tabindex="0" '
                     f'data-card-group="{ehtml(ev["group"])}" style="--class-hue:{class_hue}" '
                     f'title="Show only {ehtml(class_label)}; click again to restore" '
                     f'aria-label="Show only {ehtml(class_label)} lessons; activate again to restore the previous filter">'
                     f'<span class="class-dot" aria-hidden="true"></span>{ehtml(class_label)}</span>')
    teacher_cls = " is-missing" if fields["teacher"] == "-" else " is-alert" if fields["teacher"] in {"Andy", "Calvin"} else ""
    note_html = "".join(f' <span class="card-note">[{ehtml(note)}]</span>' for note in fields["notes"])
    helper_html = (
        f'<span class="erb-sep">&middot;</span><span class="erb-helper">Helper: {ehtml(fields["helper"])}</span>'
        if fields["helper"] else ""
    )
    return (f'<div class="chip erb-compact {st} cat-{ev["category"]} grp-{ev["group"]}{red_cls}{layer_cls}{comparison_cls}" tabindex="0" role="button" '
            f'data-date="{ehtml(ev["date"])}" data-status="{ehtml(st)}" data-cat="{ehtml(ev["category_label"])}" data-group="{ehtml(ev["group"])}" data-group-label="{ehtml(ev["group_label"])}" data-text="{ehtml(ev["text"])}" data-html="{ehtml(full_html)}"{layer_attrs}{comparison_attrs}>'
            f'{comparison_badge}<span class="status" aria-label="{ehtml(st)}">{mark}</span>{identity_html}'
            f'<div class="erb-meta"><span class="erb-location">{ehtml(fields["location"])}</span><span class="erb-sep">&middot;</span>'
            f'<span class="erb-teacher{teacher_cls}">Teacher: {ehtml(fields["teacher"])}</span>{helper_html}</div>'
            f'<div class="erb-course">{ehtml(fields["course_name"])}</div>'
            f'<div class="erb-foot"><span class="erb-time">{ehtml(fields["time"])}</span><span class="erb-sep">&middot;</span>'
            f'<span class="erb-lesson">{ehtml(fields["lesson"])}{note_html}</span></div>'
            f'<div class="fulltxt">{full_html}</div></div>')


def render_day_events(day_events):
    slot_order = ["morning", "afternoon", "night", "other"]
    groups = {slot: [] for slot in slot_order}
    for ev in sorted(day_events, key=event_sort_key):
        groups[event_slot(ev)].append(ev)
    rendered = []
    for slot in slot_order:
        group = groups[slot]
        if not group:
            continue
        if len(group) == 1:
            content = chip(group[0])
        else:
            ordered = sorted(group, key=lambda ev: (0 if event_layer(ev) == "mine" else 1, event_sort_key(ev)))
            content = '<div class="overlap-group overlap-active" data-overlap-count="{}">{}</div>'.format(
                len(ordered), ''.join(chip(ev) for ev in ordered)
            )
        rendered.append(f'<div class="time-slot slot-{slot}" data-slot="{slot}">{content}</div>')
    return ''.join(rendered)

def month_html(year, month):
    cal = calendar.Calendar(firstweekday=6)
    cells = [f'<div class="dow">{d}</div>' for d in ["Sun","Mon","Tue","Wed","Thu","Fri","Sat"]]
    for week in cal.monthdatescalendar(year, month):
        for day in week:
            ds = day.isoformat()
            evs = [] if day.month != month else by_date.get(ds, [])
            cls = "cell" + (" out" if day.month != month else "") + (" wknd" if day.weekday() >= 5 and day.month == month else "") + (" has" if evs else "") + (" holiday-cell" if any(e["category"] == "holiday" for e in evs) else "")
            cell_id = f"d-{ds}" if day.month == month else f"d-out-m{month}-{ds}"
            mon = calendar.month_abbr[day.month]
            weekday = day.strftime("%a")
            cells.append(f'<div class="{cls}" id="{cell_id}"><div class="dnum"><span class="dmon">{mon}</span><span class="dday">{day.day}</span><span class="day-hours" data-day-hours hidden></span><span class="dweekday">{weekday}</span></div>' + render_day_events(evs) + '</div>')
    grid = '<div class="gridwrap"><div class="grid">' + ''.join(cells) + '</div></div>'
    daykeys = [datetime.date(year, month, day) for day in range(1, calendar.monthrange(year, month)[1] + 1)]
    agenda_bits = []
    for day in daykeys:
        ds = day.isoformat()
        mon = calendar.month_abbr[day.month]
        day_events = by_date.get(ds, [])
        empty_class = " empty-day" if not day_events else ""
        agenda_bits.append(f'<div class="aday{empty_class}" id="a-d-{ds}" data-date="{ds}"><div class="adate"><span class="adow">{day.strftime("%a")}</span><span class="amon">{mon}</span><span class="anum">{day.day}</span><span class="day-hours" data-day-hours hidden></span></div><div class="achips">' + render_day_events(day_events) + '</div></div>')
    agenda = ''.join(agenda_bits)
    return f'<section class="month" id="m{month}"><h2>{calendar.month_name[month]} {year}</h2>{grid}<div class="agenda">{agenda}</div></section>'

counts = {"confirmed": 0, "unconfirmed": 0, "note": 0}
for e in display_events:
    counts[e['status']] = counts.get(e['status'], 0) + 1
cat_counts = {}
for e in display_events:
    cat_counts[e['category_label']] = cat_counts.get(e['category_label'], 0) + 1
layer_counts = {"mine": 0, "class": 0, "other": 0}
for e in display_events:
    layer_counts[event_layer(e)] += 1
months_html = ''.join(month_html(YEAR, m) for m in range(5, 13))
def calendar_filter_button(group):
    label, slug, group_status, first_date = group
    count = sum(1 for event in display_events if event["group"] == slug)
    return (
        f'<button class="filter course-filter {ehtml(group_status)}" data-filter="{ehtml(slug)}" '
        f'data-first-date="{ehtml(first_date)}" data-status-summary="{ehtml(group_status)}" '
        f'title="{ehtml(label)} · {ehtml(group_status)}">{ehtml(label)} ({count})</button>'
    )


calendar_filter_buckets = {"erb": [], "sen": [], "other": []}
for group in GROUPS:
    label, slug, _group_status, _first_date = group
    categories = {event["category"] for event in display_events if event["group"] == slug}
    if categories & {"erb", "methodist"}:
        bucket = "erb"
    elif categories == {"ymca"}:
        bucket = "sen"
    else:
        bucket = "other"
    calendar_filter_buckets[bucket].append(group)

filter_groups_html = ''.join(
    f'<div class="filter-group"><div class="filter-group-label">{label}</div>'
    f'<div class="filter-group-buttons">{"".join(calendar_filter_button(group) for group in sorted(calendar_filter_buckets[key], key=lambda item: natural_key(item[0])))}</div></div>'
    for key, label in (("erb", "ERB"), ("sen", "SEN"), ("other", "Other jobs"))
)

erb_filter_groups = calendar_filter_buckets["erb"]
erb_filter_counts = {
    status: sum(group[2] == status for group in erb_filter_groups)
    for status in ("upcoming", "pending", "completed", "context")
}
filter_status_summary_html = (
    f'<div class="filter-status-summary" aria-label="ERB class filter status summary">'
    f'<span class="filter-status-total">{len(erb_filter_groups)} tracked ERB classes</span>'
    + ''.join(
        f'<span class="filter-status-pill" data-filter-status="{status}">'
        f'<span class="filter-status-swatch {status}"></span>{label} {erb_filter_counts[status]}</span>'
        for status, label in (
            ("upcoming", "Upcoming"),
            ("pending", "Pending"),
            ("completed", "Completed"),
            ("context", "Full-class context"),
        )
    )
    + '</div>'
)


def summary_group_status(events):
    statuses = {event["status"] for event in events}
    if statuses == {"confirmed"}:
        return "confirmed"
    if statuses == {"unconfirmed"}:
        return "unconfirmed"
    return "mixed"


def summary_class_meta(label, events):
    if label in UPCOMING_CLASS_META:
        return UPCOMING_CLASS_META[label]
    locations = []
    for event in events:
        location = clean_location(event.get("title"), event.get("category"))
        if location != "-" and location not in locations:
            locations.append(location)
    language = "ENG" if label.upper().startswith("HK265HG") else "CHI"
    return "ERB", " / ".join(locations[:2]) or "Location in calendar", language


def provider_meta(code):
    if code.upper().startswith("MC"):
        return "MC", "循道衞理中心"
    return "CA", "基督教勵行會"


def class_summary_button(group, section):
    label, slug, _group_status, _first_date = group
    mine_events = [
        event
        for event in display_events
        if event["group"] == slug
        and event_layer(event) == "mine"
        and event["category"] in {"erb", "methodist"}
    ]
    if not mine_events:
        return None
    dated_events = [(datetime.date.fromisoformat(event["date"]), event) for event in mine_events]
    if section == "upcoming":
        section_events = [event for day, event in dated_events if day >= UPCOMING_AS_OF]
        if not section_events:
            return None
        display_day = min(datetime.date.fromisoformat(event["date"]) for event in section_events)
        date_label = f"{calendar.month_abbr[display_day.month]} {display_day.day}"
        group_status = summary_group_status(section_events)
        status_label = {"confirmed": "CONFIRMED", "unconfirmed": "UNCONFIRMED", "mixed": "MIXED"}[group_status]
        card_classes = f"{group_status} upcoming"
    else:
        last_day = max(day for day, _event in dated_events)
        if last_day >= UPCOMING_AS_OF:
            return None
        section_events = mine_events
        display_day = last_day
        date_label = f"Ended {calendar.month_abbr[display_day.month]} {display_day.day}"
        group_status = summary_group_status(section_events)
        status_label = "COMPLETED"
        card_classes = "completed"

    code = label.split(" · ", 1)[0]
    course_name = COURSE_CHINESE_NAMES.get(code, "")
    _centre, location, language = summary_class_meta(label, section_events)
    provider_code, centre = provider_meta(code)
    provider_class = provider_code.lower()
    html_button = (
        f'<button class="class-summary-card upcoming-course course-filter {card_classes}" type="button" '
        f'data-filter="{ehtml(slug)}" data-toggle-filter="1" data-first-date="{display_day.isoformat()}" '
        f'aria-label="Filter {ehtml(label)}; {ehtml(status_label)}">'
        f'<span class="upcoming-date">{ehtml(date_label)}</span>'
        f'<span class="upcoming-course-copy"><strong>{ehtml(label)}</strong>'
        f'<span>{ehtml(course_name)}</span></span>'
        f'<span class="upcoming-language">{ehtml(language)}</span>'
        f'<span class="provider-badge provider-{ehtml(provider_class)}" title="{ehtml(centre)}" '
        f'aria-label="{ehtml(provider_code)}: {ehtml(centre)}">{ehtml(provider_code)}</span>'
        f'<span class="upcoming-place"><strong>{ehtml(centre)}</strong><span>{ehtml(location)}</span></span>'
        f'<span class="summary-status">{ehtml(status_label)}</span>'
        f'</button>'
    )
    return display_day, natural_key(label), html_button


def class_summary_section(section, heading, heading_id):
    courses = [item for item in (class_summary_button(group, section) for group in GROUPS) if item]
    courses.sort(key=lambda item: (item[0], item[1]), reverse=section == "completed")
    status_key = (
        '<div class="class-summary-key" aria-label="Class confirmation key">'
        '<span><i></i>Confirmed</span><span class="pending"><i></i>Unconfirmed</span></div>'
        if section == "upcoming" else ""
    )
    return (
        f'<section class="class-summary {section}-summary" aria-labelledby="{heading_id}">'
        f'<div class="class-summary-head"><div id="{heading_id}" class="section-h">{heading}</div>{status_key}</div>'
        f'<div class="class-summary-list">{"".join(item[2] for item in courses)}</div></section>'
    )


upcoming_summary_html = class_summary_section("upcoming", "Upcoming ERB classes", "upcomingHeading")
completed_summary_html = class_summary_section("completed", "Completed ERB classes", "completedHeading")

TIMELINE_START = datetime.date(YEAR, 5, 1)
TIMELINE_END = datetime.date(YEAR, 12, 31)
TIMELINE_DAYS = (TIMELINE_END - TIMELINE_START).days + 1


def timeline_percent(day, centered=False):
    offset = (day - TIMELINE_START).days + (0.5 if centered else 0)
    return max(0.0, min(100.0, offset / TIMELINE_DAYS * 100))


def short_date(day):
    return f"{calendar.month_abbr[day.month]} {day.day}"


span_months = []
span_track_months = []
span_month_toggles = []
for month in range(5, 13):
    month_days = calendar.monthrange(YEAR, month)[1]
    day_cells = ''.join(f'<span class="span-day">{day}</span>' for day in range(1, month_days + 1))
    span_months.append(
        f'<div class="span-month" data-span-month="{month}" style="--month-days:{month_days}">'
        f'<strong>{calendar.month_abbr[month]}</strong><div class="span-days">{day_cells}</div></div>'
    )
    span_track_months.append(
        f'<span class="span-track-month" data-span-month="{month}" style="--month-days:{month_days}"></span>'
    )
    span_month_toggles.append(
        f'<label class="span-month-toggle"><input type="checkbox" data-span-month-toggle="{month}" checked '
        f'aria-label="Show {calendar.month_name[month]}"><span>{calendar.month_abbr[month]}</span></label>'
    )

def span_identity(ev):
    if ev["category"] == "ymca":
        match = SEN_CODE_RE.search(ev["text"])
        code = match.group(1).upper() if match else "SEN"
        return f"sen-{code.lower()}", f"YMCA SEN · {code}", "特殊教育需要支援課程"
    label = ev["group_label"]
    name = next((name for code, name in COURSE_CHINESE_NAMES.items() if code in label.upper()), "")
    return ev["group"], label, name


span_group_map = {}
for ev in display_events:
    if ev["category"] not in {"erb", "methodist", "ymca"}:
        continue
    slug, label, chinese_name = span_identity(ev)
    span_group_map.setdefault(
        slug,
        {"label": label, "chinese_name": chinese_name, "events": []},
    )["events"].append(ev)

span_instances = []
for base_slug, group in span_group_map.items():
    instances = split_span_instances(group["events"])
    for index, instance_events in enumerate(instances, 1):
        first_instance_day = datetime.date.fromisoformat(min(event["date"] for event in instance_events))
        instance_slug = base_slug if len(instances) == 1 else f"{base_slug}-c{index}"
        instance_label = group["label"]
        if len(instances) > 1:
            instance_label = f'{instance_label} · {calendar.month_abbr[first_instance_day.month].upper()} {YEAR}'
        span_instances.append({
            "slug": instance_slug,
            "base_slug": base_slug,
            "label": instance_label,
            "chinese_name": group["chinese_name"],
            "events": instance_events,
        })

span_rows = []
span_controls = []
for group in sorted(
    span_instances,
    key=lambda item: (min(ev["date"] for ev in item["events"]), natural_key(item["label"])),
):
    slug = group["slug"]
    group_events = sorted(group["events"], key=lambda ev: (ev["date"], event_sort_key(ev)))
    first_day = datetime.date.fromisoformat(group_events[0]["date"])
    last_day = datetime.date.fromisoformat(group_events[-1]["date"])
    span_days = (last_day - first_day).days + 1
    bar_left = timeline_percent(first_day)
    bar_width = span_days / TIMELINE_DAYS * 100
    class_hue = zlib.crc32(group["label"].encode("utf-8")) % 360
    events_by_day = {}
    for ev in group_events:
        events_by_day.setdefault(ev["date"], []).append(ev)

    markers = []
    my_dates = 0
    for ds, day_events in sorted(events_by_day.items()):
        day = datetime.date.fromisoformat(ds)
        mine = any(event_layer(ev) == "mine" for ev in day_events)
        mine_confirmed = any(event_layer(ev) == "mine" and ev["status"] == "confirmed" for ev in day_events)
        if mine:
            my_dates += 1
        status = "unconfirmed" if any(ev["status"] == "unconfirmed" for ev in day_events) else "confirmed"
        details = []
        for ev in day_events:
            fields = event_fields(ev)
            detail = f'{fields["lesson"]} · {fields["teacher"]} · {fields["time"]}'
            if detail not in details:
                details.append(detail)
        tooltip = f'{short_date(day)}: ' + " / ".join(details)
        markers.append(
            f'<button class="span-marker {"mine" if mine else "other"} {status}" type="button" '
            f'data-date="{ds}" data-mine="{1 if mine else 0}" data-mine-confirmed="{1 if mine_confirmed else 0}" '
            f'data-group-label="{ehtml(group["label"])}" '
            f'data-details="{ehtml(tooltip)}" title="{ehtml(tooltip)}" '
            f'aria-label="{ehtml(tooltip)}; {"Garett teaches" if mine else "other tutor or TBC"}"></button>'
        )

    range_label = f"{short_date(first_day)}–{short_date(last_day)}"
    lifecycle = lifecycle_status(group_events)
    class_type = "sen" if all(ev["category"] == "ymca" for ev in group_events) else "erb"
    span_controls.append((slug, group["label"], range_label, lifecycle, class_type))
    span_rows.append(
        f'<div class="span-row" data-span-group="{ehtml(slug)}" data-base-group="{ehtml(group["base_slug"])}" data-first="{first_day.isoformat()}" '
        f'data-last="{last_day.isoformat()}" data-lesson-dates="{len(events_by_day)}" data-my-dates="{my_dates}">'
        f'<div class="span-label"><strong>{ehtml(group["label"])}</strong>'
        f'<span class="span-course-name">{ehtml(group["chinese_name"])}</span>'
        f'<small>{ehtml(range_label)} · My dates {my_dates}/{len(events_by_day)}</small></div>'
        f'<div class="span-track"><div class="span-track-grid">{"".join(span_track_months)}</div>'
        f'<div class="span-bar status-{lifecycle}" style="--span-hue:{class_hue}" '
        f'data-first-label="{ehtml(short_date(first_day))}" data-last-label="{ehtml(short_date(last_day))}" '
        f'title="{ehtml(group["label"])} · {ehtml(range_label)}">'
        f'<span class="span-bar-label">{ehtml(group["label"])}</span>{"".join(markers)}</div></div></div>'
    )

span_timeline_html = (
    '<div class="span-scroll" tabindex="0" aria-label="Scrollable class timeline from May to December 2026">'
    '<div class="span-table"><div class="span-axis"><div class="span-axis-label">Course / class</div>'
    f'<div class="span-axis-track">{"".join(span_months)}</div></div>{"".join(span_rows)}</div></div>'
    if span_rows else '<div class="span-empty">No tracked classes.</div>'
)
span_course_toggles = ''.join(
    f'<label class="span-course-toggle status-{lifecycle} type-{class_type}"><input type="checkbox" data-span-course="{ehtml(slug)}" checked>'
    f'<span><strong>{ehtml(label)}</strong><small>{ehtml(range_label)}</small>'
    f'<span class="span-status-tag">{ehtml("SEN" if class_type == "sen" else lifecycle)}</span></span></label>'
    for slug, label, range_label, lifecycle, class_type in span_controls
)
span_erb_count = sum(class_type == "erb" for _slug, _label, _range, _status, class_type in span_controls)
span_sen_count = sum(class_type == "sen" for _slug, _label, _range, _status, class_type in span_controls)
span_course_picker = (
    f'<section id="spanCoursePicker" class="span-course-picker" aria-label="Class visibility">'
    f'<div class="span-course-picker-head"><strong>Class visibility</strong><span id="spanCourseCount">{len(span_controls)}/{len(span_controls)} ON</span>'
    f'<span class="span-course-breakdown">{len(span_controls)} total = {span_erb_count} ERB + {span_sen_count} SEN</span>'
    f'<div class="span-course-actions"><button type="button" data-span-course-action="all">All on</button>'
    f'<button type="button" data-span-course-action="none">All off</button></div></div>'
    f'<div class="span-course-picker-body"><div class="span-course-options">{span_course_toggles}</div></div></section>'
)
span_month_controls = (
    '<div class="span-month-controls" role="group" aria-label="Visible months">'
    '<span class="span-month-controls-label">Months</span>'
    + ''.join(span_month_toggles)
    + '</div>'
)
erb_code_legend = ''.join(
    '<section class="course-family-card">'
    f'<div class="course-family-name">{ehtml(family["name"])}</div>'
    '<div class="course-family-members">'
    + ''.join(
        f'<span class="course-family-member"><b>{ehtml(code)}</b><span>{ehtml(centre)}</span>'
        f'{f"<em>{ehtml(note)}</em>" if note else ""}</span>'
        for code, centre, note in family["members"]
    )
    + '</div></section>'
    for family in COURSE_FAMILIES
)
version_items = json.loads(VERSIONS_SRC.read_text(encoding="utf-8"))
current_version = next((item for item in version_items if item.get("latest")), None)
if not current_version:
    raise ValueError("versions.json has no latest timetable version")
version_rows = ''.join(
    f'<button class="version-menu-item{" current" if item["id"] == current_version["id"] else ""}" type="button" '
    f'data-version-id="{ehtml(item["id"])}"><strong>{ehtml(item["label"])}</strong>'
    f'<span>{ehtml(item["summary"])}</span>'
    f'{"<span class=\"version-menu-badge\">Current</span>" if item["id"] == current_version["id"] else ""}</button>'
    for item in version_items
)
version_selector_html = (
    f'<details id="topVersionSelector" class="version-menu"><summary>'
    f'<span class="version-menu-kicker">Versions</span><span class="version-menu-current">{ehtml(current_version["label"].split(" - ")[-1])}</span>'
    f'<span class="version-menu-summary">{ehtml(current_version["summary"])}</span><span class="version-menu-arrow" aria-hidden="true">&#9662;</span>'
    f'</summary><div class="version-menu-list" role="list">{version_rows}</div></details>'
)

HTML = f'''<!doctype html><html lang="en"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1, minimum-scale=1, maximum-scale=6, user-scalable=yes">
<title>Garett's ERB — Super Timetable</title>
<meta name="description" content="ERB / YMCA / school teaching timetable, May to December 2026. Solid frame = confirmed; dotted frame = unconfirmed.">
<link rel="apple-touch-icon" sizes="180x180" href="icon-180.png"><link rel="icon" type="image/png" sizes="32x32" href="favicon-32.png"><link rel="icon" type="image/png" sizes="192x192" href="icon-192.png"><link rel="manifest" href="manifest.webmanifest">
<meta name="apple-mobile-web-app-capable" content="yes"><meta name="mobile-web-app-capable" content="yes"><meta name="apple-mobile-web-app-title" content="Garett's ERB"><meta name="application-name" content="Garett's ERB"><meta name="apple-mobile-web-app-status-bar-style" content="default"><meta name="theme-color" content="#0f7074">
<meta name="erb-build" content="{BUILD_ID}">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<script>window.ERB_BUILD_ID='{BUILD_ID}';(function(){{if(!/^https?:$/.test(location.protocol))return;var p=new URLSearchParams(location.search);if(p.get('build')!==window.ERB_BUILD_ID){{p.set('build',window.ERB_BUILD_ID);location.replace(location.pathname+'?'+p.toString()+location.hash);}}}})();</script>
<style>{CSS}</style></head><body><main class="wrap">
<div class="hero"><div><h1 class="title"><span class="y">ERB</span> Super Timetable</h1><p class="sub">May–December 2026 · personal timetable plus complete ERB class context · solid frame = confirmed, dotted frame = unconfirmed</p></div><div class="actions"><a class="btn" href="#today" id="todayBtn">Today</a><a class="btn" href="#m5">May</a><a class="btn" href="#m6">Jun</a><a class="btn" href="#m7">Jul</a><a class="btn" href="#m8">Aug</a><a class="btn" href="#m9">Sep</a><a class="btn" href="#m10">Oct</a><a class="btn" href="#m11">Nov</a><a class="btn" href="#m12">Dec</a></div></div>
{version_selector_html}
<div id="viewTabs" class="view-tabs" role="tablist" aria-label="Timetable layout"><button id="calendarTab" class="view-tab active" type="button" role="tab" aria-selected="true" aria-controls="calendarView" data-view="calendar">Calendar</button><button id="spansTab" class="view-tab" type="button" role="tab" aria-selected="false" aria-controls="spansView" data-view="spans">Class spans</button></div>
<section id="calendarView" class="view-panel" role="tabpanel" aria-labelledby="calendarTab">
<div class="stats"><div class="stat"><b>{len(display_events)}</b> total entries</div><div class="stat"><b>{layer_counts['mine']}</b> my schedule</div><div class="stat"><b>{layer_counts['class']}</b> other class lessons</div><div class="stat"><b>{counts.get('confirmed',0)}</b> confirmed</div><div class="stat"><b>{counts.get('unconfirmed',0)}</b> unconfirmed</div></div>
<div class="legend"><div class="legend-card"><span class="sample confirmed"></span> Confirmed / 已確認</div><div class="legend-card"><span class="sample unconfirmed"></span> Unconfirmed / 未確認</div><div class="legend-card"><span class="sample class-layer"></span> Full class context</div>{comparison_legend_html}<div class="legend-card"><span class="sample note"></span> Note / holiday</div></div>
<div id="transitNotice" class="transit-notice" hidden>Transit reminders appear between your lesson slots. Red means less than 30 minutes remains after travel, so there is no reliable meal break.</div>
<div class="course-code-heading"><div class="section-h">ERB course families</div><span class="course-code-count">{len(COURSE_FAMILIES)} course families</span></div><div class="course-code-legend">{erb_code_legend}</div>
{upcoming_summary_html}
{completed_summary_html}
<div class="filter-heading"><div id="filterArea" class="section-h filter-jump-target">Filter by course / class</div><div class="filter-master"><button class="filter course-filter active" data-filter="all">All ({len(display_events)})</button>{comparison_filter_html}</div></div>{filter_status_summary_html}<div class="filter-groups">{filter_groups_html}</div>
{months_html}
<div class="foot">Sources: <b>{ehtml(SRC.name)}</b>, <b>{ehtml(OVERRIDES_SRC.name)}</b>, and <b>{ehtml(CONTEXT_SRC.name)}</b>. The supplemental layer never overwrites a workbook entry. Generated from Excel border styles: solid/medium = confirmed, dashed = unconfirmed.</div>
</section><section id="spansView" class="view-panel" role="tabpanel" aria-labelledby="spansTab" hidden>
<div id="spanShell" class="span-shell"><div class="span-head"><h2>Class spans</h2><p>Each row is one class. Every visible grid column is one calendar day; darker markers are Garett's lesson dates.</p></div><div class="span-control-bar"><div class="span-tools"><div class="span-mode-switch" role="group" aria-label="Class span lesson filter"><button class="span-mode-option" type="button" data-mode="mine-confirmed">ME CONF</button><button class="span-mode-option" type="button" data-mode="mine-all">ME ALL</button><button class="span-mode-option active" type="button" data-mode="both">ALL FULL</button></div><button id="spanLabelsToggle" class="span-tool-button" type="button" aria-pressed="true">Course names ON</button><div class="span-zoom" role="group" aria-label="Class span timeline zoom"><button id="spanZoomOut" type="button" aria-label="Zoom timeline out" title="Zoom timeline out">−</button><button id="spanZoomReset" class="span-zoom-value" type="button" aria-label="Reset timeline zoom" title="Reset timeline zoom">100%</button><button id="spanZoomIn" type="button" aria-label="Zoom timeline in" title="Zoom timeline in">+</button></div>{span_course_picker}{span_month_controls}<div class="span-legend"><span class="span-legend-item"><span class="span-key"></span>Garett teaches</span><span class="span-legend-item"><span class="span-key other"></span>Other tutor / TBC</span></div></div></div>{span_timeline_html}</div>
</section>
</main><div id="modeSwitch" class="floating-mode-switch" role="group" aria-label="Timetable view and navigation"><button id="floatingToday" class="today-option" type="button" aria-label="Go to today" title="Go to today"><span class="mode-main">TODAY</span></button><button id="floatingTop" class="top-option" type="button" aria-label="Back to course filters" title="Back to course filters"><span class="mode-main" aria-hidden="true">&uarr;</span><span class="mode-sub">FILTER</span></button><button id="floatingVersions" class="version-option" type="button" aria-label="Back to version selector" title="Back to version selector"><span class="mode-main">VER</span></button><button class="mode-option" type="button" data-mode="mine-confirmed" aria-label="Me: confirmed lessons" title="Me: confirmed lessons"><span class="mode-main">ME</span><span class="mode-sub">CONF</span></button><button class="mode-option" type="button" data-mode="mine-all" aria-label="Me: confirmed and unconfirmed lessons" title="Me: confirmed and unconfirmed lessons"><span class="mode-main">ME</span><span class="mode-sub">ALL</span></button><button class="mode-option active" type="button" data-mode="both" aria-label="All full timetable and clear course filter" title="All full timetable and clear course filter"><span class="mode-main">ALL</span><span class="mode-sub">FULL</span></button></div><div id="modal" class="modal" hidden><div class="modal-card"><button class="modal-x" aria-label="Close">×</button><div class="modal-h"></div><div class="modal-date"></div><div class="modal-body"></div></div></div>
<script>
if('serviceWorker' in navigator&&/^https?:$/.test(location.protocol)){{window.addEventListener('load',()=>navigator.serviceWorker.register('./sw.js?build='+window.ERB_BUILD_ID).then(r=>r.update()).catch(()=>{{}}));}}
const primaryTabs=Array.from(document.querySelectorAll('.view-tab'));
const primaryPanels=Array.from(document.querySelectorAll('.view-panel'));
const primaryViewScroll={{calendar:0,spans:0}};
let activePrimaryView=null;
function selectPrimaryView(view,focusTab=false){{
  const selected=primaryTabs.find(tab=>tab.dataset.view===view)||primaryTabs[0];
  const nextView=selected.dataset.view;
  const previousView=activePrimaryView;
  if(previousView&&previousView!==nextView) primaryViewScroll[previousView]=window.scrollY;
  primaryTabs.forEach(tab=>{{const active=tab===selected;tab.classList.toggle('active',active);tab.setAttribute('aria-selected',String(active));tab.tabIndex=active?0:-1;}});
  primaryPanels.forEach(panel=>panel.hidden=panel.id!==selected.getAttribute('aria-controls'));
  document.body.classList.toggle('span-view-active',nextView==='spans');
  activePrimaryView=nextView;
  window.__activePrimaryView=nextView;
  const params=new URLSearchParams(location.search);
  if(nextView==='spans') params.set('view','spans'); else params.delete('view');
  history.replaceState(null,'',location.pathname+(params.toString()?'?'+params.toString():'')+location.hash);
  if(previousView&&previousView!==nextView) requestAnimationFrame(()=>window.scrollTo(0,primaryViewScroll[nextView]||0));
  if(focusTab) selected.focus();
}}
primaryTabs.forEach((tab,index)=>{{
  tab.addEventListener('click',()=>selectPrimaryView(tab.dataset.view));
  tab.addEventListener('keydown',event=>{{if(!['ArrowLeft','ArrowRight'].includes(event.key))return;event.preventDefault();const offset=event.key==='ArrowRight'?1:-1;const next=primaryTabs[(index+offset+primaryTabs.length)%primaryTabs.length];selectPrimaryView(next.dataset.view,true);}});
}});
selectPrimaryView(new URLSearchParams(location.search).get('view')==='spans'?'spans':'calendar');
const modal=document.getElementById('modal');
function openChip(el){{
  const st=el.dataset.status, cat=el.dataset.cat, txt=el.dataset.text, html=el.dataset.html||'', date=el.dataset.date, source=el.dataset.source||'', layer=el.dataset.layer;
  const changed=el.dataset.changed==='1', previous=el.dataset.previous||'', previousStatus=el.dataset.previousStatus||'', changeKind=el.dataset.changeKind||'';
  const esc=value=>String(value||'').replace(/[&<>"']/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[c]));
  modal.querySelector('.modal-h').textContent=cat;
  modal.querySelector('.modal-date').innerHTML=date+' · <span class="pill '+st+'">'+(st==='confirmed'?'Confirmed / 已確認':st==='unconfirmed'?'Unconfirmed / 未確認':'Note / 備註')+'</span>'+(layer==='class'?' <span class="pill class-layer">Full class context</span>':'')+(changed?' <span class="pill changed-pill">Changed in {COMPARE_LABEL}</span>':'');
  const oldContent=changeKind==='new'?'<span class="comparison-new">Not present in {COMPARE_BASELINE_LABEL}</span>':esc(previous)+(previousStatus?'<div class="old-status">Status in {COMPARE_BASELINE_LABEL}: '+esc(previousStatus)+'</div>':'');
  const comparison=changed?'<div class="comparison-old"><strong>{COMPARE_BASELINE_LABEL} content</strong><div>'+oldContent+'</div></div>':'';
  modal.querySelector('.modal-body').innerHTML='<span class="modal-current-label">Current {COMPARE_LABEL} content</span><div class="modal-current">'+(html||txt)+'</div>'+comparison+(source?'<div class="modal-source">Source: '+esc(source)+'</div>':'');
  modal.hidden=false;
}}
document.querySelectorAll('.chip').forEach(el=>{{el.addEventListener('click',()=>openChip(el));el.addEventListener('keydown',e=>{{if(e.key==='Enter'||e.key===' '){{e.preventDefault();openChip(el)}}}})}});
function openSpanMarker(el){{
  modal.querySelector('.modal-h').textContent=el.dataset.groupLabel||'Class lesson';
  modal.querySelector('.modal-date').textContent=el.dataset.date||'';
  modal.querySelector('.modal-body').textContent=el.dataset.details||'';
  modal.hidden=false;
}}
document.querySelectorAll('.span-marker').forEach(el=>el.addEventListener('click',()=>openSpanMarker(el)));
modal.querySelector('.modal-x').onclick=()=>modal.hidden=true; modal.addEventListener('click',e=>{{if(e.target===modal) modal.hidden=true}}); document.addEventListener('keydown',e=>{{if(e.key==='Escape') modal.hidden=true}});
function isPortraitAgenda(){{return window.matchMedia('(orientation: portrait) and (max-width: 820px)').matches;}}
function jumpToFilter(btn){{
  const f=btn.dataset.filter;
  const candidates=Array.from(document.querySelectorAll(f==='changed'?'.chip[data-changed="1"]':'.chip.grp-'+f)).filter(ch=>ch.style.display!=='none');
  const preferred=candidates.find(ch=>isPortraitAgenda()?ch.closest('.agenda'):ch.closest('.grid'))||candidates[0];
  const ds=preferred&&preferred.dataset.date;
  if(!ds) return;
  const target=document.getElementById((isPortraitAgenda()?'a-d-':'d-')+ds)||document.getElementById('a-d-'+ds)||document.getElementById('d-'+ds);
  if(!target) return;
  setTimeout(()=>{{
    if(isPortraitAgenda()){{
      const y=window.pageYOffset + target.getBoundingClientRect().top - 10;
      window.scrollTo(0, Math.max(0, y));
    }} else {{
      target.scrollIntoView({{block:'center', inline:'center', behavior:'auto'}});
    }}
  }}, 40);
}}
window.__courseFilter='all';
window.__layerMode='both';
window.__cardCourseFilter=null;
window.__upcomingFilterState=null;
function syncCourseFilterUI(){{
  document.querySelectorAll('.course-filter').forEach(btn=>btn.classList.toggle('active',btn.dataset.filter===window.__courseFilter));
  const activeGroup=window.__cardCourseFilter&&window.__cardCourseFilter.group;
  document.querySelectorAll('.card-course-filter').forEach(tag=>tag.classList.toggle('card-filter-active',tag.dataset.cardGroup===activeGroup));
}}
function syncLayerModeUI(){{
  document.querySelectorAll('.mode-option,.span-mode-option').forEach(option=>option.classList.toggle('active',option.dataset.mode===window.__layerMode));
}}
const spanShell=document.getElementById('spanShell');
const spanCourseInputs=Array.from(document.querySelectorAll('[data-span-course]'));
const spanCourseCount=document.getElementById('spanCourseCount');
function syncSpanCourseCount(){{
  const enabled=spanCourseInputs.filter(input=>input.checked).length;
  if(spanCourseCount) spanCourseCount.textContent=enabled+'/'+spanCourseInputs.length+' ON';
}}
let spanLabelsVisible=true;
const spanLabelsToggle=document.getElementById('spanLabelsToggle');
function syncSpanLabels(){{
  spanShell.classList.toggle('hide-span-labels',!spanLabelsVisible);
  spanLabelsToggle.setAttribute('aria-pressed',String(spanLabelsVisible));
  spanLabelsToggle.textContent='Course names '+(spanLabelsVisible?'ON':'OFF');
}}
const spanZoomLevels=[8,12,16,22,30,40];
let spanZoomIndex=2;
const spanMonthInputs=Array.from(document.querySelectorAll('[data-span-month-toggle]'));
const spanMonthDays={{5:31,6:30,7:31,8:31,9:30,10:31,11:30,12:31}};
function visibleSpanMonths(){{return spanMonthInputs.filter(input=>input.checked).map(input=>Number(input.dataset.spanMonthToggle));}}
function spanMonthOffset(month,visibleMonths){{let offset=0;visibleMonths.forEach(value=>{{if(value<month) offset+=spanMonthDays[value];}});return offset;}}
function layoutSpanTimeline(){{
  const visibleMonths=visibleSpanMonths();
  const dayWidth=spanZoomLevels[spanZoomIndex];
  const visibleDays=visibleMonths.reduce((sum,month)=>sum+spanMonthDays[month],0);
  spanShell.style.setProperty('--span-day-width',dayWidth+'px');
  spanShell.style.setProperty('--span-timeline-width',(visibleDays*dayWidth)+'px');
  document.querySelectorAll('[data-span-month]').forEach(element=>{{element.hidden=!visibleMonths.includes(Number(element.dataset.spanMonth));}});
  document.querySelectorAll('.span-row').forEach(row=>{{
    const firstParts=row.dataset.first.split('-').map(Number);
    const lastParts=row.dataset.last.split('-').map(Number);
    let startIndex=Infinity;
    let endIndex=-Infinity;
    visibleMonths.forEach(month=>{{
      if(month<firstParts[1]||month>lastParts[1]) return;
      const firstDay=month===firstParts[1]?firstParts[2]:1;
      const lastDay=month===lastParts[1]?lastParts[2]:spanMonthDays[month];
      const monthOffset=spanMonthOffset(month,visibleMonths);
      startIndex=Math.min(startIndex,monthOffset+firstDay-1);
      endIndex=Math.max(endIndex,monthOffset+lastDay);
    }});
    const bar=row.querySelector('.span-bar');
    const rangeVisible=Number.isFinite(startIndex)&&Number.isFinite(endIndex)&&endIndex>startIndex;
    row.dataset.monthHidden=rangeVisible?'0':'1';
    bar.hidden=!rangeVisible;
    if(!rangeVisible) return;
    const barLeft=startIndex*dayWidth;
    bar.style.left=barLeft+'px';
    bar.style.width=Math.max(dayWidth,(endIndex-startIndex)*dayWidth)+'px';
    row.querySelectorAll('.span-marker').forEach(marker=>{{
      const parts=marker.dataset.date.split('-').map(Number);
      const month=parts[1];
      const monthVisible=visibleMonths.includes(month);
      marker.dataset.monthVisible=monthVisible?'1':'0';
      if(monthVisible){{
        const center=(spanMonthOffset(month,visibleMonths)+parts[2]-.5)*dayWidth;
        marker.style.left=(center-barLeft)+'px';
      }}
    }});
  }});
}}
function spanLayoutMetrics(){{
  if(window.matchMedia('(orientation: landscape) and (max-height: 700px) and (max-width: 1400px) and (pointer: coarse)').matches) return {{label:205}};
  if(window.matchMedia('(max-width: 820px)').matches) return {{label:230}};
  return {{label:280}};
}}
function applySpanZoom(preservePosition=false){{
  const previousMax=Math.max(0,document.documentElement.scrollWidth-innerWidth);
  const previousRatio=previousMax?scrollX/previousMax:0;
  const metrics=spanLayoutMetrics();
  spanShell.style.setProperty('--span-label-width',(spanLabelsVisible?metrics.label:0)+'px');
  layoutSpanTimeline();
  document.getElementById('spanZoomReset').textContent=Math.round(spanZoomLevels[spanZoomIndex]/16*100)+'%';
  document.getElementById('spanZoomOut').disabled=spanZoomIndex===0;
  document.getElementById('spanZoomIn').disabled=spanZoomIndex===spanZoomLevels.length-1;
  if(preservePosition) requestAnimationFrame(()=>{{
    const nextMax=Math.max(0,document.documentElement.scrollWidth-innerWidth);
    window.scrollTo(previousRatio*nextMax,scrollY);
  }});
}}
spanLabelsToggle.addEventListener('click',()=>{{spanLabelsVisible=!spanLabelsVisible;syncSpanLabels();applySpanZoom(true);}});
document.getElementById('spanZoomOut').addEventListener('click',()=>{{if(spanZoomIndex>0){{spanZoomIndex-=1;applySpanZoom(true);}}}});
document.getElementById('spanZoomIn').addEventListener('click',()=>{{if(spanZoomIndex<spanZoomLevels.length-1){{spanZoomIndex+=1;applySpanZoom(true);}}}});
document.getElementById('spanZoomReset').addEventListener('click',()=>{{spanZoomIndex=2;applySpanZoom(true);}});
spanMonthInputs.forEach(input=>input.addEventListener('change',()=>{{
  if(!input.checked&&visibleSpanMonths().length===0){{input.checked=true;return;}}
  applySpanZoom(true);
  applySpanFilters();
}}));
spanCourseInputs.forEach(input=>input.addEventListener('change',()=>{{syncSpanCourseCount();applySpanFilters();}}));
document.querySelectorAll('[data-span-course-action]').forEach(button=>button.addEventListener('click',()=>{{
  const checked=button.dataset.spanCourseAction==='all';
  spanCourseInputs.forEach(input=>input.checked=checked);
  syncSpanCourseCount();
  applySpanFilters();
}}));
let spanResizeTimer=null;
window.addEventListener('resize',()=>{{clearTimeout(spanResizeTimer);spanResizeTimer=setTimeout(()=>applySpanZoom(),100);}},{{passive:true}});
function applySpanFilters(){{
  const mode=window.__layerMode;
  document.querySelectorAll('.span-row').forEach(row=>{{
    const courseInput=spanCourseInputs.find(input=>input.dataset.spanCourse===row.dataset.spanGroup);
    const courseEnabled=!courseInput||courseInput.checked;
    let visible=0;
    row.querySelectorAll('.span-marker').forEach(marker=>{{
      const modeMatch=mode==='both'||(mode==='mine-all'&&marker.dataset.mine==='1')||(mode==='mine-confirmed'&&marker.dataset.mineConfirmed==='1');
      const show=modeMatch&&marker.dataset.monthVisible!=='0';
      marker.hidden=!show;
      if(show) visible+=1;
    }});
    row.dataset.courseHidden=courseEnabled?'0':'1';
    row.hidden=!courseEnabled||row.dataset.monthHidden==='1'||visible===0;
  }});
}}
const transitMinutes={{
  'sheung_shui|four_seas':64,'four_seas|sheung_shui':64,
  'four_seas|choi_wan':40,'choi_wan|four_seas':40,
  'ymca_yau_ma_tei|choi_wan':37,'choi_wan|ymca_yau_ma_tei':37,
  'ymca_yau_ma_tei|four_seas':12,'four_seas|ymca_yau_ma_tei':12,
  'shun_tin|choi_wan':30,'choi_wan|shun_tin':30,
  'sheung_shui|choi_wan':65,'choi_wan|sheung_shui':65
}};
const centreLabels={{sheung_shui:'Sheung Shui',four_seas:'Four Seas',choi_wan:'Choi Wan',ymca_yau_ma_tei:'YMCA',wan_chai:'Wan Chai',lam_tin:'Lam Tin',shun_tin:'Shun Tin'}};
function visibleMineChips(slot){{
  return Array.from(slot.querySelectorAll('.chip')).filter(ch=>ch.style.display!=='none'&&ch.dataset.layer==='mine'&&ch.dataset.start&&ch.dataset.end&&ch.dataset.centre);
}}
function refreshTransitBars(){{
  document.querySelectorAll('.transit-bar').forEach(bar=>bar.remove());
  const notice=document.getElementById('transitNotice');
  const personal=window.__layerMode==='mine-confirmed'||window.__layerMode==='mine-all';
  notice.hidden=!personal;
  if(!personal) return;
  document.querySelectorAll('.cell,.achips').forEach(container=>{{
    const slots=Array.from(container.querySelectorAll(':scope > .time-slot')).filter(slot=>!slot.classList.contains('slot-hidden'));
    for(let index=0;index<slots.length-1;index+=1){{
      const fromSlot=slots[index],toSlot=slots[index+1];
      const fromChips=visibleMineChips(fromSlot),toChips=visibleMineChips(toSlot);
      if(!fromChips.length||!toChips.length) continue;
      const fromEnd=Math.max(...fromChips.map(ch=>Number(ch.dataset.end)));
      const toStart=Math.min(...toChips.map(ch=>Number(ch.dataset.start)));
      const gap=toStart-fromEnd;
      if(gap<=0) continue;
      const routes=[];
      const seen=new Set();
      fromChips.filter(ch=>Number(ch.dataset.end)===fromEnd).forEach(from=>toChips.filter(ch=>Number(ch.dataset.start)===toStart).forEach(to=>{{
        const a=from.dataset.centre,b=to.dataset.centre,key=a+'|'+b;
        if(seen.has(key)) return;
        seen.add(key);
        const travel=a===b?0:transitMinutes[key];
        if(travel===undefined) return;
        routes.push({{a,b,travel,spare:gap-travel}});
      }}));
      if(!routes.length) continue;
      routes.sort((a,b)=>a.spare-b.spare||b.travel-a.travel);
      const route=routes[0],tight=route.spare<30;
      const bar=document.createElement('div');
      bar.className='transit-bar'+(tight?' tight':'');
      const routeLabel=route.a===route.b?'Same centre':(centreLabels[route.a]||route.a)+' to '+(centreLabels[route.b]||route.b);
      const timing=route.spare<0?Math.abs(route.spare)+'m short':route.spare+'m spare';
      bar.innerHTML='<strong>'+routeLabel+'</strong> · '+route.travel+'m transit · '+timing+(tight?' · NO MEAL BUFFER':'');
      fromSlot.after(bar);
    }}
  }});
}}
function parseTeachingIntervals(value){{
  return String(value||'').split(',').filter(Boolean).map(part=>part.split('-').map(Number)).filter(interval=>interval.length===2&&Number.isFinite(interval[0])&&Number.isFinite(interval[1])&&interval[1]>interval[0]);
}}
function mergeTeachingMinutes(intervals){{
  if(!intervals.length) return 0;
  intervals.sort((a,b)=>a[0]-b[0]||a[1]-b[1]);
  const merged=[];
  intervals.forEach(interval=>{{
    const last=merged[merged.length-1];
    if(last&&interval[0]<=last[1]) last[1]=Math.max(last[1],interval[1]);
    else merged.push([interval[0],interval[1]]);
  }});
  return merged.reduce((total,interval)=>total+interval[1]-interval[0],0);
}}
function formatTeachingMinutes(minutes){{
  const hours=Math.floor(minutes/60),remainder=minutes%60;
  if(hours&&remainder) return hours+'h'+String(remainder).padStart(2,'0');
  if(hours) return hours+'h';
  return remainder+'m';
}}
function refreshDailyHours(){{
  const mode=window.__layerMode;
  const personal=mode==='mine-confirmed'||mode==='mine-all';
  document.querySelectorAll('[data-day-hours]').forEach(badge=>{{
    badge.hidden=true;
    badge.textContent='';
    badge.removeAttribute('title');
    badge.removeAttribute('aria-label');
    if(!personal) return;
    const container=badge.closest('.cell,.aday');
    if(!container) return;
    const intervals=[];
    container.querySelectorAll('.chip[data-layer="mine"]').forEach(chip=>{{
      const status=chip.dataset.status;
      const include=mode==='mine-confirmed'?status==='confirmed':(status==='confirmed'||status==='unconfirmed');
      if(include) intervals.push(...parseTeachingIntervals(chip.dataset.teachingIntervals));
    }});
    const minutes=mergeTeachingMinutes(intervals);
    if(!minutes) return;
    const label=formatTeachingMinutes(minutes);
    badge.textContent=label;
    badge.title=(mode==='mine-confirmed'?'Confirmed':'Confirmed and unconfirmed')+' teaching time; travel excluded';
    badge.setAttribute('aria-label',label+' teaching time, travel excluded');
    badge.hidden=false;
  }});
}}
function applyFilters(){{
  const f=window.__courseFilter, mode=window.__layerMode;
  window.__filterActive = f !== 'all';
  document.querySelectorAll('.chip').forEach(ch=>{{
    const courseMatch=f==='all'||(f==='changed'?ch.dataset.changed==='1':ch.classList.contains('grp-'+f));
    const isMike=ch.dataset.cat==='Mike Sir';
    const isHoliday=ch.dataset.cat==='Holiday';
    const isMine=ch.dataset.layer==='mine';
    const layerMatch=mode==='both'||(mode==='mine-all'&&(isMine||isMike))||(mode==='mine-confirmed'&&(isMike||isHoliday||(isMine&&ch.dataset.status==='confirmed')));
    ch.style.display=courseMatch&&layerMatch?'':'none';
  }});
  document.querySelectorAll('.overlap-group').forEach(group=>{{
    const visible=Array.from(group.querySelectorAll('.chip')).filter(ch=>ch.style.display!=='none').length;
    group.classList.toggle('overlap-active',visible>=2);
  }});
  document.querySelectorAll('.time-slot').forEach(slot=>{{
    const visible=Array.from(slot.querySelectorAll('.chip')).some(ch=>ch.style.display!=='none');
    slot.classList.toggle('slot-hidden',!visible);
  }});
  document.querySelectorAll('.cell').forEach(cell=>{{ const visible=Array.from(cell.querySelectorAll('.chip')).some(ch=>ch.style.display!=='none'); if(cell.querySelector('.chip')) cell.classList.toggle('has', visible); }});
  document.querySelectorAll('.aday').forEach(day=>{{ const chips=Array.from(day.querySelectorAll('.chip')); if(chips.length) day.style.display=chips.some(ch=>ch.style.display!=='none')?'':'none'; }});
  refreshDailyHours();
  refreshTransitBars();
  applySpanFilters();
}}
document.querySelectorAll('.course-filter').forEach(btn=>btn.addEventListener('click',()=>{{
  const requested=btn.dataset.filter;
  let f=requested;
  if(btn.dataset.toggleFilter==='1'){{
    const state=window.__upcomingFilterState;
    if(state&&state.group===requested){{
      f=state.previousFilter;
      window.__upcomingFilterState=null;
    }} else if(state){{
      state.group=requested;
    }} else {{
      window.__upcomingFilterState={{group:requested,previousFilter:window.__courseFilter}};
    }}
  }} else {{
    window.__upcomingFilterState=null;
  }}
  window.__cardCourseFilter=null;
  window.__courseFilter=f;
  syncCourseFilterUI();
  applyFilters();
  if(f!=='all') jumpToFilter(btn);
}}));
function captureCardAnchor(tag){{
  const target=tag.closest(isPortraitAgenda()?'.aday':'.cell')||tag.closest('.aday')||tag.closest('.cell');
  return target?{{id:target.id,top:target.getBoundingClientRect().top}}:null;
}}
function toggleCardCourseFilter(tag){{
  const group=tag.dataset.cardGroup;
  const anchor=captureCardAnchor(tag);
  const current=window.__cardCourseFilter;
  if(current&&current.group===group){{
    window.__courseFilter=current.previousFilter;
    window.__cardCourseFilter=null;
  }} else if(current){{
    current.group=group;
    window.__courseFilter=group;
  }} else {{
    window.__cardCourseFilter={{group,previousFilter:window.__courseFilter}};
    window.__courseFilter=group;
  }}
  syncCourseFilterUI();
  window.__restoringModeAnchor=true;
  applyFilters();
  restoreModeAnchor(anchor);
}}
document.querySelectorAll('.card-course-filter').forEach(tag=>{{
  tag.addEventListener('click',event=>{{event.stopPropagation();toggleCardCourseFilter(tag);}});
  tag.addEventListener('keydown',event=>{{
    if(event.key==='Enter'||event.key===' '){{event.preventDefault();event.stopPropagation();toggleCardCourseFilter(tag);}}
  }});
}});
function captureModeAnchor(){{
  const selector=isPortraitAgenda()?'.agenda .aday':'.grid .cell:not(.out)';
  const candidates=Array.from(document.querySelectorAll(selector)).filter(el=>{{
    const style=getComputedStyle(el), rect=el.getBoundingClientRect();
    return style.display!=='none'&&rect.bottom>0&&rect.top<innerHeight;
  }});
  if(!candidates.length) return null;
  const anchorY=innerHeight/2;
  const distance=el=>{{const rect=el.getBoundingClientRect();return rect.top<=anchorY&&rect.bottom>=anchorY?0:Math.min(Math.abs(rect.top-anchorY),Math.abs(rect.bottom-anchorY));}};
  const target=candidates.reduce((best,el)=>distance(el)<distance(best)?el:best);
  return {{id:target.id,top:target.getBoundingClientRect().top}};
}}
function restoreModeAnchor(anchor){{
  if(!anchor){{window.__restoringModeAnchor=false;return;}}
  requestAnimationFrame(()=>setTimeout(()=>{{
    const target=document.getElementById(anchor.id);
    if(!target){{window.__restoringModeAnchor=false;return;}}
    if(isPortraitAgenda()){{target.classList.remove('empty-day');target.style.display='';}}
    const delta=target.getBoundingClientRect().top-anchor.top;
    const previousScrollBehavior=document.documentElement.style.scrollBehavior;
    document.documentElement.style.scrollBehavior='auto';
    window.__restoringModeAnchor=true;
    window.scrollTo(0,Math.max(0,window.scrollY+delta));
    document.documentElement.style.scrollBehavior=previousScrollBehavior;
    setTimeout(()=>{{window.__restoringModeAnchor=false;}},80);
  }},30));
}}
window.__modeCompareAnchor=null;
window.__restoringModeAnchor=false;
window.addEventListener('scroll',()=>{{if(!window.__restoringModeAnchor) window.__modeCompareAnchor=null;}},{{passive:true}});
document.getElementById('floatingTop').addEventListener('click',()=>{{window.__modeCompareAnchor=null;document.getElementById('filterArea').scrollIntoView({{block:'start',behavior:'smooth'}});}});
document.getElementById('floatingVersions').addEventListener('click',()=>{{
  window.__modeCompareAnchor=null;
  window.scrollTo({{top:0,left:0,behavior:'smooth'}});
}});
document.querySelectorAll('.version-menu-item').forEach(btn=>btn.addEventListener('click',()=>{{
  const root=location.pathname.includes('/versions/')?'../../':'./';
  location.assign(root+'versions/'+encodeURIComponent(btn.dataset.versionId)+'/?v=redtext1');
}}));
document.querySelectorAll('.mode-option').forEach(btn=>btn.addEventListener('click',()=>{{
  const anchor=window.__modeCompareAnchor||captureModeAnchor();
  window.__modeCompareAnchor=anchor;
  window.__layerMode=btn.dataset.mode;
  if(btn.dataset.mode==='both'){{
    window.__courseFilter='all';
    window.__upcomingFilterState=null;
    window.__cardCourseFilter=null;
    syncCourseFilterUI();
  }}
  syncLayerModeUI();
  window.__restoringModeAnchor=true;
  applyFilters();
  restoreModeAnchor(anchor);
}}));
document.querySelectorAll('.span-mode-option').forEach(btn=>btn.addEventListener('click',()=>{{
  window.__layerMode=btn.dataset.mode;
  window.__modeCompareAnchor=null;
  syncLayerModeUI();
  applyFilters();
}}));
syncCourseFilterUI();
syncLayerModeUI();
syncSpanLabels();
syncSpanCourseCount();
applySpanZoom();
applyFilters();
(function(){{
 const pad=n=>String(n).padStart(2,'0');
 const localDate=d=>`${{d.getFullYear()}}-${{pad(d.getMonth()+1)}}-${{pad(d.getDate())}}`;
 const params=new URLSearchParams(location.search);
 const override=params.get('today');
 const ds=/^\\d{{4}}-\\d{{2}}-\\d{{2}}$/.test(override||'') ? override : localDate(new Date());
 const gridToday=document.getElementById('d-'+ds);
 const agendaToday=document.getElementById('a-d-'+ds);
 if(gridToday) gridToday.classList.add('today');
 if(agendaToday){{agendaToday.classList.add('today');agendaToday.classList.remove('empty-day');}}
 const todayBtn=document.getElementById('todayBtn');
 function isPortraitAgenda(){{return window.matchMedia('(orientation: portrait) and (max-width: 820px)').matches;}}
 function focusToday(force=false){{
   if(window.__filterActive&&!force) return;
   const portrait=isPortraitAgenda();
   const target=portrait ? agendaToday : gridToday;
   if(!target){{todayBtn.href='#m6'; return;}}
   todayBtn.href='#';
   if(portrait){{
     agendaToday.style.display='';
     const y=window.pageYOffset + target.getBoundingClientRect().top - 10;
     const previousScrollBehavior=document.documentElement.style.scrollBehavior;
     document.documentElement.style.scrollBehavior='auto';
     window.scrollTo(0, Math.max(0, y));
     document.documentElement.style.scrollBehavior=previousScrollBehavior;
   }} else {{
     target.scrollIntoView({{block:'center', inline:'center', behavior:'auto'}});
   }}
 }}
 window.focusTimetableToday=()=>focusToday(true);
 document.getElementById('floatingToday').addEventListener('click',()=>{{window.__modeCompareAnchor=null;window.focusTimetableToday();}});
 todayBtn.addEventListener('click', e=>{{e.preventDefault(); window.__filterActive=false; window.__modeCompareAnchor=null; focusToday();}});
 focusToday();
 requestAnimationFrame(focusToday);
 window.addEventListener('load', ()=>setTimeout(focusToday,80));
 window.addEventListener('orientationchange', ()=>setTimeout(focusToday,450));
 setTimeout(focusToday,300);
 setTimeout(focusToday,900);
}})();
</script></body></html>'''

SW = f'''const BUILD_ID = '{BUILD_ID}';

self.addEventListener('install', event => {{
  self.skipWaiting();
}});

self.addEventListener('activate', event => {{
  event.waitUntil(
    caches.keys()
      .then(keys => Promise.all(keys.map(key => caches.delete(key))))
      .then(() => self.clients.claim())
  );
}});

self.addEventListener('fetch', event => {{
  const url = new URL(event.request.url);
  if (url.origin !== self.location.origin) return;
  if (event.request.method !== 'GET') return;
  const req = new Request(event.request, {{ cache: 'no-store' }});
  event.respondWith(fetch(req).catch(() => fetch(event.request)));
}});
'''

(OUTDIR / 'index.html').write_text(HTML, encoding='utf-8')
(OUTDIR / '.nojekyll').write_text('', encoding='utf-8')
(OUTDIR / 'sw.js').write_text(SW, encoding='utf-8')
(OUTDIR / 'events.json').write_text(json.dumps(events, ensure_ascii=False, indent=2), encoding='utf-8')
(OUTDIR / 'summary.json').write_text(json.dumps({"source": str(SRC), "override_source": str(OVERRIDES_SRC), "override_revision": override_revision, "override_confirmation": override_confirmation, "events": len(events), "display_events": len(display_events), "context_events": len(context_events), "class_spans": len(span_rows), "comparison_baseline": str(COMPARE_BASELINE), "comparison_label": COMPARE_LABEL, "changed_in_version": len(changed_events), "counts": counts, "layers": layer_counts, "categories": cat_counts, "months": MONTH_SHEETS}, ensure_ascii=False, indent=2), encoding='utf-8')
(OUTDIR / 'manifest.webmanifest').write_text(json.dumps({"id":"./","name":"Garett's ERB","short_name":"Garett's ERB","description":"Garett's ERB teaching timetable","start_url":"./?v=redtext1&build=" + BUILD_ID,"scope":"./","display":"standalone","background_color":"#eef1f6","theme_color":"#0f7074","icons":[{"src":"icon-192.png","sizes":"192x192","type":"image/png","purpose":"any maskable"},{"src":"icon-512.png","sizes":"512x512","type":"image/png","purpose":"any maskable"}]}, ensure_ascii=False, indent=2), encoding='utf-8')
try:
    from PIL import Image, ImageDraw, ImageFont
    def make_icon(size, filename):
        output_path = OUTDIR / filename
        img = Image.new('RGB', (size, size), '#0f7074')
        d = ImageDraw.Draw(img)
        pad = max(2, size // 12)
        radius = max(4, size // 6)
        d.rounded_rectangle([pad, pad, size-pad, size-pad], radius=radius, fill='#ffffff')
        d.rounded_rectangle([pad, pad, size-pad, size//3], radius=radius, fill='#f2a33a')
        d.rectangle([pad, size//5, size-pad, size//3], fill='#f2a33a')
        font_path = r'C:/Windows/Fonts/segoeuib.ttf'
        g_font = ImageFont.truetype(font_path, max(9, int(size*.37)))
        erb_font = ImageFont.truetype(font_path, max(6, int(size*.16)))
        d.text((size//2, int(size*.54)), 'G', anchor='mm', font=g_font, fill='#0f7074')
        d.text((size//2, int(size*.80)), 'ERB', anchor='mm', font=erb_font, fill='#1d2734')
        img.save(output_path)
    make_icon(32, 'favicon-32.png')
    make_icon(180, 'icon-180.png')
    make_icon(192, 'icon-192.png')
    make_icon(512, 'icon-512.png')
except Exception as e:
    print('icon generation skipped', e)
print('generated', OUTDIR)
print('events', len(events), counts)
for k, v in sorted(cat_counts.items()):
    print(k, v)
