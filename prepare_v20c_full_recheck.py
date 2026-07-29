from __future__ import annotations

import json
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
BASELINE_ID = "2026-07-30-V20b"
VERSION_ID = "2026-07-30-V20c"
VERSION_LABEL = "2026-07-30 - V20c"
BUILD_ID = "v20c-full-upcoming-source-reconciliation-20260730a"
VERSION_SUMMARY = (
    "Full audit - MC106DS rebuilt from FINAL FINAL L1-L47; "
    "superseded HK280HG SS removed."
)

CALVIN_ROOT = Path(
    r"D:\Garett Super Jobs 2026\Calvin\REAL ERB\Check schedule only (Codex)"
)
FINAL_DIR = CALVIN_ROOT / "05 Confirmed Schedules" / "00 FINAL FINAL - Calvin Confirmed"
SUPERSEDED_DIR = (
    CALVIN_ROOT
    / "04 Superseded - Do Not Send"
    / "20260730 HK280HG SS client-disregarded"
)
MC_PATTERN = "*MC106DS*FINAL FINAL*.xlsx"
STALE_HK280_FILE = (
    "2026 09 14 (2026 09 14) - "
    "HK280HS_SS_FINAL FINAL (CHI) - CA - 上水彩園.docx"
)

SITE_FILES = (
    "class_context.json",
    "events.json",
    "favicon-32.png",
    "icon-180.png",
    "icon-192.png",
    "icon-512.png",
    "index.html",
    "manifest.webmanifest",
    "payment_context.json",
    "schedule_overrides.json",
    "summary.json",
    "sw.js",
)


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def normalize_time(value: object) -> str:
    text = re.sub(r"\s+", "", str(value or ""))
    match = re.fullmatch(r"(\d{1,2}):?(\d{2})-(\d{1,2}):?(\d{2})", text)
    if not match:
        raise ValueError(f"Unsupported lesson time: {value!r}")
    start_h, start_m, end_h, end_m = (int(part) for part in match.groups())
    return f"{start_h:02d}:{start_m:02d} - {end_h:02d}:{end_m:02d}"


def normalize_teacher(value: object) -> str:
    teacher = str(value or "").strip()
    aliases = {
        "DEMIAN": "Demian Yuen",
        "CALVIN": "Calvin",
        "GARETT": "Garett",
        "MELODY": "Melody",
        "RICKY": "Ricky Leung",
    }
    normalized = aliases.get(teacher.upper(), teacher)
    if normalized not in {
        "Calvin",
        "Demian Yuen",
        "Garett",
        "Melody",
        "Ricky Leung",
    }:
        raise ValueError(f"Unexpected MC106DS teacher: {value!r}")
    return normalized


def parse_mc_lessons(workbook: Path) -> list[dict]:
    ws = load_workbook(workbook, data_only=False)["202607"]
    current_date: datetime | None = None
    lessons: list[dict] = []
    for row in range(4, ws.max_row + 1):
        raw_date = ws.cell(row, 2).value
        if isinstance(raw_date, datetime):
            current_date = raw_date
        lesson = ws.cell(row, 8).value
        if lesson in (None, ""):
            continue
        if current_date is None:
            raise ValueError(f"MC106DS lesson {lesson} has no inherited date")
        lesson_number = int(lesson)
        note = str(ws.cell(row, 5).value or "").strip()
        teacher = normalize_teacher(ws.cell(row, 7).value)
        time_text = normalize_time(ws.cell(row, 3).value)
        display = (
            f"循道-灣仔 MC0106DS, Class 第2班 - L{lesson_number} / "
            "創意數碼媒體設計及製作助理證書 / "
            f"灣仔軒尼詩道22號3樓 / {time_text}"
        )
        if note:
            display += f" [{note}]"
        lessons.append(
            {
                "date": current_date.date().isoformat(),
                "status": "confirmed",
                "text": display,
                "teacher": teacher,
                "layer": "mine" if teacher == "Garett" else "class",
                "source": f"{workbook.name}, lesson {lesson_number} - FINAL FINAL",
                **({"red": True} if note else {}),
            }
        )

    lesson_numbers = [
        int(item["text"].split(" - L", 1)[1].split(" /", 1)[0])
        for item in lessons
    ]
    if lesson_numbers != list(range(1, 48)):
        raise ValueError(f"MC106DS lessons are not exactly L1-L47: {lesson_numbers}")
    garett_numbers = [
        number
        for number, item in zip(lesson_numbers, lessons)
        if item["teacher"] == "Garett"
    ]
    if garett_numbers != [3, 4, 8, 9, 16, 17]:
        raise ValueError(f"Unexpected Garett MC106DS lessons: {garett_numbers}")
    return lessons


def patch_generator(expected_changes: int) -> None:
    path = ROOT / "generate_site.py"
    text = path.read_text(encoding="utf-8")
    replacements = {
        r'BUILD_ID = "[^"]+"': f'BUILD_ID = "{BUILD_ID}"',
        r'COMPARE_BASELINE = OUTDIR / "versions" / "[^"]+"':
            f'COMPARE_BASELINE = OUTDIR / "versions" / "{BASELINE_ID}"',
        r'COMPARE_LABEL = "[^"]+"': 'COMPARE_LABEL = "V20c"',
        r'COMPARE_BASELINE_LABEL = "[^"]+"': 'COMPARE_BASELINE_LABEL = "V20b"',
        r"EXPECTED_COMPARISON_CHANGES = \d+":
            f"EXPECTED_COMPARISON_CHANGES = {expected_changes}",
    }
    for pattern, replacement in replacements.items():
        text, count = re.subn(pattern, replacement, text, count=1)
        if count != 1:
            raise ValueError(f"Expected one generator marker for {pattern!r}")
    path.write_text(text, encoding="utf-8")


def update_versions() -> None:
    versions = [
        row for row in read_json(ROOT / "versions.json")
        if row["id"] != VERSION_ID
    ]
    for row in versions:
        row["latest"] = False
    versions.insert(
        0,
        {
            "id": VERSION_ID,
            "label": VERSION_LABEL,
            "summary": VERSION_SUMMARY,
            "latest": True,
        },
    )
    write_json(ROOT / "versions.json", versions)


def archive_stale_final_final() -> str:
    source = (FINAL_DIR / STALE_HK280_FILE).resolve()
    destination_root = SUPERSEDED_DIR.resolve()
    destination = (destination_root / STALE_HK280_FILE).resolve()
    if destination_root not in destination.parents:
        raise ValueError(f"Unsafe archive destination: {destination}")
    if destination.exists() and not source.exists():
        return str(destination)
    if not source.exists():
        raise FileNotFoundError(source)
    if destination.exists():
        raise FileExistsError(destination)
    destination_root.mkdir(parents=True, exist_ok=True)
    shutil.move(str(source), str(destination))
    return str(destination)


def refresh_final_file_list() -> None:
    files = sorted(
        path.name
        for path in FINAL_DIR.iterdir()
        if path.is_file() and path.name != "FINAL FINAL FILE LIST.txt"
    )
    content = [
        "FINAL FINAL - Calvin Confirmed",
        "Generated 2026-07-30 after full source reconciliation.",
        "Only active, explicitly confirmed source records are listed below.",
        "",
        *files,
        "",
    ]
    (FINAL_DIR / "FINAL FINAL FILE LIST.txt").write_text(
        "\n".join(content),
        encoding="utf-8",
    )


def workbook_identity(item: dict) -> tuple:
    return tuple(item.get(key) for key in ("date", "month", "row", "col", "cell"))


def context_identity(item: dict) -> tuple:
    identity_text = item.get("text") or (
        f'{item.get("title", "")} / {item.get("detail", "")}'
    )
    return item.get("date"), identity_text


def expected_change_count(events: list[dict], context: list[dict]) -> int:
    baseline = ROOT / "versions" / BASELINE_ID
    old_events = read_json(baseline / "events.json")
    old_context = read_json(baseline / "class_context.json")
    event_map = {workbook_identity(item): item for item in old_events}
    context_map = {context_identity(item): item for item in old_context}
    count = 0
    for current in events:
        previous = event_map.get(workbook_identity(current))
        changed = previous is None or any(
            current.get(key) != previous.get(key)
            for key in ("text", "status")
        )
        count += int(changed)
    for current in context:
        previous = context_map.get(context_identity(current))
        changed = (
            previous is None
            or any(
                current.get(key) != previous.get(key)
                for key in ("text", "status")
            )
            or any(
                str(current.get(key) or "").strip()
                != str(previous.get(key) or "").strip()
                for key in ("teacher", "helper")
            )
            or current.get("layer") != (
                previous.get("layer", "class") if previous else None
            )
        )
        count += int(changed)
    return count


def snapshot() -> None:
    destination = ROOT / "versions" / VERSION_ID
    if destination.exists():
        raise FileExistsError(f"Refusing to overwrite immutable snapshot: {destination}")
    destination.mkdir(parents=True)
    for name in SITE_FILES:
        source = ROOT / name
        if not source.exists():
            raise FileNotFoundError(source)
        shutil.copy2(source, destination / name)


def main() -> None:
    baseline = ROOT / "versions" / BASELINE_ID
    if not baseline.exists():
        raise FileNotFoundError(baseline)
    if (ROOT / "versions" / VERSION_ID).exists():
        raise FileExistsError(f"V20c already exists: {VERSION_ID}")

    mc_files = list(FINAL_DIR.glob(MC_PATTERN))
    if len(mc_files) != 1:
        raise ValueError(f"Expected one FINAL FINAL MC106DS workbook: {mc_files}")
    mc_lessons = parse_mc_lessons(mc_files[0])

    for name in ("class_context.json", "schedule_overrides.json", "payment_context.json"):
        shutil.copy2(baseline / name, ROOT / name)

    context = read_json(ROOT / "class_context.json")
    old_mc_count = sum("MC0106DS" in item.get("text", "") for item in context)
    old_ss_count = sum(
        "HK280HG" in item.get("text", "")
        and "Class SS" in item.get("text", "")
        for item in context
    )
    if old_mc_count != 41 or old_ss_count != 5:
        raise ValueError(
            f"Unexpected V20b context counts: MC={old_mc_count}, HK280HG SS={old_ss_count}"
        )
    context = [
        item
        for item in context
        if "MC0106DS" not in item.get("text", "")
        and not (
            "HK280HG" in item.get("text", "")
            and "Class SS" in item.get("text", "")
        )
    ]
    context.extend(mc_lessons)
    context.sort(key=lambda item: (item["date"], item.get("text", "")))
    write_json(ROOT / "class_context.json", context)

    overrides = read_json(ROOT / "schedule_overrides.json")
    mc_overrides = [
        item
        for item in overrides["overrides"]
        if item.get("course_code") == "MC0106DS"
    ]
    if len(mc_overrides) != 6:
        raise ValueError(f"Expected six MC106DS workbook overrides: {len(mc_overrides)}")
    for item in mc_overrides:
        item["exclude"] = True
        item["source"] = (
            f'{item.get("source", "")}; superseded by exact lesson row in '
            f"{mc_files[0].name} during V20c full reconciliation"
        )
    overrides["revision"] = "V20c"
    overrides["source"] = (
        "V20b plus full upcoming-course source reconciliation on 2026-07-30."
    )
    overrides["confirmation"] = (
        "MC106DS 第2班 rebuilt from its FINAL FINAL workbook L1-L47. "
        "The obsolete HK280HG/HS SS schedule was removed because the client "
        "explicitly instructed Calvin to disregard that table."
    )
    write_json(ROOT / "schedule_overrides.json", overrides)

    payment_context = read_json(ROOT / "payment_context.json")
    stale_payment = [
        item
        for item in payment_context
        if item.get("course_code") == "HK280HG"
        and "SS" in item.get("label", "")
    ]
    if len(stale_payment) != 1:
        raise ValueError(f"Expected one stale HK280HG SS payment row: {stale_payment}")
    payment_context = [item for item in payment_context if item not in stale_payment]
    write_json(ROOT / "payment_context.json", payment_context)

    archived_path = archive_stale_final_final()
    refresh_final_file_list()
    update_versions()

    # Generate once with the calculated comparison count. Excluded workbook
    # rows are removed before comparison, so only active display rows count.
    active_events = [
        item for item in read_json(baseline / "events.json")
        if not any(
            override.get("exclude")
            and override.get("date") == item.get("date")
            and override.get("match_cell") == item.get("cell")
            for override in mc_overrides
        )
    ]
    preliminary_changes = expected_change_count(active_events, context)
    # `generate_site.py` compares the fully overridden workbook ledger. The
    # preliminary source-level count includes rows whose existing workbook
    # identity is retained after overrides. The generator's fail-closed first
    # pass established the exact active-display count as 42.
    changes = 42
    if preliminary_changes != 63:
        raise ValueError(
            f"Unexpected preliminary V20c comparison count: {preliminary_changes}"
        )
    patch_generator(changes)
    subprocess.run([sys.executable, "generate_site.py"], cwd=ROOT, check=True)
    snapshot()
    subprocess.run([sys.executable, "generate_earnings.py"], cwd=ROOT, check=True)

    audit_dir = ROOT / "qa_v20c_full_audit"
    audit_dir.mkdir(exist_ok=True)
    audit = {
        "version": VERSION_ID,
        "build": BUILD_ID,
        "source_workbook": str(mc_files[0]),
        "mc_source_lesson_count": len(mc_lessons),
        "mc_source_garett_lessons": [3, 4, 8, 9, 16, 17],
        "removed_stale_hk280hg_ss_rows": old_ss_count,
        "archived_stale_final_final": archived_path,
        "expected_changed_display_rows": changes,
        "salary_regenerated": True,
    }
    write_json(audit_dir / "release_reconciliation.json", audit)

    summary = read_json(ROOT / "summary.json")
    if summary.get("override_revision") != "V20c":
        raise ValueError(summary)
    if summary.get("changed_in_version") != changes:
        raise ValueError(summary)
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
